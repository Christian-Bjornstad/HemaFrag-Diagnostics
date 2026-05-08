from __future__ import annotations

import os
import shutil
import sys
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook, load_workbook

from config import APP_SETTINGS
from core.batch import generate_jobs, run_batch_jobs


DATA_ROOT = Path("/Volumes/T7 Shield/DATA/2025_data")
OUTPUT_ROOT = Path("/Volumes/T7 Shield/HemaFrag_2025_safe_reruns_2026-04-28")
STATUS_WORKBOOK = OUTPUT_ROOT / "2025_monthly_run_status.xlsx"
YEAR = "2025"


def month_dirs() -> dict[str, list[Path]]:
    grouped: dict[str, list[Path]] = defaultdict(list)
    for path in sorted(DATA_ROOT.iterdir()):
        if not path.is_dir():
            continue
        name = path.name
        if not name.startswith(f"{YEAR}_"):
            continue
        parts = name.split("_")
        if len(parts) < 2:
            continue
        month = parts[1]
        if len(month) == 2 and month.isdigit():
            grouped[month].append(path)
    return dict(sorted(grouped.items()))


def count_nonempty_fsa(paths: list[Path]) -> int:
    total = 0
    for folder in paths:
        for fsa in folder.glob("*.fsa"):
            try:
                if fsa.stat().st_size > 0:
                    total += 1
            except OSError:
                continue
    return total


def ensure_status_workbook() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    if STATUS_WORKBOOK.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Status"
    ws.append(
        [
            "Month",
            "StartedAt",
            "FinishedAt",
            "Folders",
            "NonEmptyFsa",
            "Jobs",
            "Status",
            "OutputFolder",
            "TrackingWorkbook",
            "Notes",
        ]
    )
    wb.save(STATUS_WORKBOOK)


def load_status_map() -> dict[str, dict[str, str]]:
    ensure_status_workbook()
    wb = load_workbook(STATUS_WORKBOOK, data_only=True)
    ws = wb["Monthly Status"]
    headers = [str(ws.cell(row=1, column=col).value or "") for col in range(1, ws.max_column + 1)]
    out: dict[str, dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        month = str(ws.cell(row=row, column=1).value or "").strip()
        if not month:
            continue
        values: dict[str, str] = {}
        for col, header in enumerate(headers[1:], start=2):
            value = ws.cell(row=row, column=col).value
            values[header] = "" if value is None else str(value)
        out[month] = values
    return out


def upsert_status(
    month: str,
    *,
    started_at: str | None = None,
    finished_at: str | None = None,
    folders: int | None = None,
    nonempty_fsa: int | None = None,
    jobs: int | None = None,
    status: str | None = None,
    output_folder: str | None = None,
    tracking_workbook: str | None = None,
    notes: str | None = None,
) -> None:
    ensure_status_workbook()
    wb = load_workbook(STATUS_WORKBOOK)
    ws = wb["Monthly Status"]

    target_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "") == month:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=month)

    values = [
        started_at,
        finished_at,
        folders,
        nonempty_fsa,
        jobs,
        status,
        output_folder,
        tracking_workbook,
        notes,
    ]
    for idx, value in enumerate(values, start=2):
        if value is not None:
            ws.cell(row=target_row, column=idx, value=value)

    wb.save(STATUS_WORKBOOK)


def resolve_tracking_workbook(month_output_root: Path) -> Path | None:
    reports_dir = month_output_root / f"reports_{datetime.now().date().isoformat()}"
    candidates = [
        reports_dir / "Clonality_Tracking.xlsx",
        reports_dir / "HemaFrag_QC_Trends.xlsx",
        month_output_root / "Clonality_Tracking.xlsx",
        month_output_root / "HemaFrag_QC_Trends.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def configure_safe_settings() -> None:
    os.environ["HEMAFRAG_SKIP_DEEP_SEARCH"] = "True"
    APP_SETTINGS["active_analysis"] = "clonality"
    APP_SETTINGS.setdefault("engine", {})
    APP_SETTINGS["engine"]["use_rust"] = True
    APP_SETTINGS["engine"]["rust_worker_pool_size"] = 1
    APP_SETTINGS["engine"]["rust_timeout_seconds"] = 45
    APP_SETTINGS["engine"]["rust_timeout_seconds_rox"] = 60
    APP_SETTINGS["engine"]["rust_timeout_seconds_liz"] = 45


def should_skip_month(month: str) -> bool:
    status_map = load_status_map()
    row = status_map.get(month)
    if not row:
        return False
    return row.get("Status", "") == "completed" and bool(row.get("TrackingWorkbook", "").strip())


def reset_partial_month_output(month_output: Path) -> None:
    if not month_output.exists():
        return
    for child_name in ("ASSAY_REPORTS", f"reports_{datetime.now().date().isoformat()}"):
        child = month_output / child_name
        if child.exists():
            shutil.rmtree(child, ignore_errors=True)
    for file_name in ("Clonality_Tracking.xlsx", "HemaFrag_QC_Trends.xlsx"):
        candidate = month_output / file_name
        if candidate.exists():
            try:
                candidate.unlink()
            except OSError:
                pass


def run_month(month: str, folders: list[Path]) -> None:
    started = datetime.now().isoformat(timespec="seconds")
    nonempty = count_nonempty_fsa(folders)
    month_output = OUTPUT_ROOT / f"{YEAR}_{month}"
    month_output.mkdir(parents=True, exist_ok=True)

    if should_skip_month(month):
        return

    reset_partial_month_output(month_output)

    upsert_status(
        month,
        started_at=started,
        folders=len(folders),
        nonempty_fsa=nonempty,
        status="running",
        output_folder=str(month_output),
        notes="starting",
    )

    if nonempty == 0:
        upsert_status(
            month,
            finished_at=datetime.now().isoformat(timespec="seconds"),
            jobs=0,
            status="skipped_empty",
            tracking_workbook="",
            notes="No non-empty .fsa files found.",
        )
        return

    jobs = generate_jobs(folders, aggregate_patients=True)
    upsert_status(month, jobs=len(jobs), notes="jobs generated")

    if not jobs:
        upsert_status(
            month,
            finished_at=datetime.now().isoformat(timespec="seconds"),
            status="skipped_no_jobs",
            tracking_workbook="",
            notes="No usable jobs generated from month folders.",
        )
        return

    try:
        result = run_batch_jobs(
            jobs=jobs,
            output_base=month_output,
            out_folder_tmpl="ASSAY_REPORTS",
            outfile_html_tmpl="QC_REPORT_{name}.html",
            excel_name_tmpl="HemaFrag_QC_Trends.xlsx",
            pipeline_scope="all",
            assay_filter="",
            aggregate_dit_reports=True,
            continue_on_error=True,
            max_workers=1,
        )
        tracking = resolve_tracking_workbook(month_output)
        status = "completed"
        notes = f"result_keys={sorted(result.keys())}"
        if tracking is None:
            status = "completed_no_workbook"
            notes = "Run returned but no tracking workbook was found."
        upsert_status(
            month,
            finished_at=datetime.now().isoformat(timespec="seconds"),
            status=status,
            tracking_workbook=str(tracking) if tracking else "",
            notes=notes,
        )
    except Exception as exc:
        err = "".join(traceback.format_exception_only(type(exc), exc)).strip()
        upsert_status(
            month,
            finished_at=datetime.now().isoformat(timespec="seconds"),
            status="failed",
            tracking_workbook="",
            notes=err,
        )


def main() -> None:
    configure_safe_settings()
    ensure_status_workbook()
    grouped = month_dirs()
    for month in [f"{m:02d}" for m in range(1, 13)]:
        run_month(month, grouped.get(month, []))


if __name__ == "__main__":
    main()
