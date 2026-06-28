"""TabClonalityInterpretation — Phase 1 / T-1.1 of Plan 11.

Render a per-row comparison of rule suggestion vs. ML suggestion, color
coded. Off by default: only loads if `interpretation.enabled = True` in
APP_SETTINGS. With ML disabled it still works as a rule-only review surface.

The widget is intentionally self-contained (no core.analyses.clonality
imports at module load) so headless tests don't need a working pipeline.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush, QColor, QPalette
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

# Color tokens — match gui_qt/styles.py palette. Local copy so the tab is
# headless-testable without depending on QSS module load.
_COLOR_AGREE = QColor("#10b981")  # green
_COLOR_DISAGREE = QColor("#f59e0b")  # amber
_COLOR_FORCE_REVIEW = QColor("#ef4444")  # red
_COLOR_OFF = QColor("#475569")  # slate-600


def _classify_row(entry: dict[str, Any], ml_present: bool) -> str:
    """Return one of {'agree', 'disagree', 'force_review'}."""
    rule_label = str(entry.get("ClonalitySuggestion") or "").strip()
    ml_label = str(entry.get("ClonalityMLSuggestion") or "").strip()
    review = bool(entry.get("ClonalityReviewNeeded"))
    if rule_label == "usikker_review" or review:
        return "force_review"
    if not ml_present:
        return "agree"  # rule-only mode; treat as agreement (green)
    if rule_label and ml_label and rule_label != ml_label:
        return "disagree"
    return "agree"


def _color_for(kind: str) -> QColor:
    if kind == "force_review":
        return _COLOR_FORCE_REVIEW
    if kind == "disagree":
        return _COLOR_DISAGREE
    return _COLOR_AGREE


class TabClonalityInterpretation(QWidget):
    """Compare rule vs ML per-row in a colored table.

    Public surface for tests/usages:
        load_from_entries(entries): main entry point (takes a list
            of entry dicts from `tracking_excel` or pipeline).
        load_batch_from_tracking(folder): load the most recent
            `reports_<date>/Clonality_Tracking.xlsx`.
        set_inline_synth_entries(): 6-row demo when no data is
            available (used by tests).
    """

    HEADERS = (
        "DIT",
        "Assay",
        "Group",
        "Rule",
        "ML",
        "Conf",
        "Review?",
        "Dom peak (bp)",
        "Evidence",
    )

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._rows: list[dict[str, Any]] = []
        self._ml_present = False
        self._build_ui()

    # ---- UI construction ------------------------------------------------

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        # Toolbar
        toolbar = QHBoxLayout()
        title = QLabel("Clonality Interpretation — Rule vs. ML Comparison")
        title.setObjectName("PageTitle")
        toolbar.addWidget(title)
        toolbar.addStretch()

        self._batch_combo = QComboBox()
        self._batch_combo.setMinimumWidth(220)
        self._batch_combo.addItem("(no batch loaded)")
        toolbar.addWidget(self._batch_combo)

        self._disagreements_only = QCheckBox("Show only disagreements")
        self._disagreements_only.toggled.connect(self._refresh_table)
        toolbar.addWidget(self._disagreements_only)

        self._refresh_btn = QPushButton("Refresh")
        self._refresh_btn.clicked.connect(self._refresh_btn_clicked)
        toolbar.addWidget(self._refresh_btn)
        layout.addLayout(toolbar)

        # Table
        self._table = QTableWidget(0, len(self.HEADERS), self)
        self._table.setHorizontalHeaderLabels(self.HEADERS)
        header = self._table.horizontalHeader()
        for i in range(len(self.HEADERS)):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(len(self.HEADERS) - 1, QHeaderView.ResizeMode.Stretch)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        layout.addWidget(self._table, stretch=1)

        # Status bar
        self._status_label = QLabel("No batch loaded. Use Refresh after a clonality batch run.")
        self._status_label.setObjectName("StatusBarText")
        layout.addWidget(self._status_label)

    # ---- API ------------------------------------------------------------

    def load_from_entries(self, entries: list[dict[str, Any]] | None = None) -> None:
        if entries is None:
            entries = []
        self._rows = list(entries)
        self._ml_present = any(
            "ClonalityMLSuggestion" in e for e in self._rows
        )
        self._refresh_table()
        self._refresh_status()

    def load_batch_from_tracking(self, tracking_file: Path | str | None = None) -> int:
        """Load entries from a `Clonality_Tracking.xlsx` if available.

        Returns the count of entries loaded (0 if no file found).
        Falls back to inline synth entries when no batch path is provided
        and none is on disk.
        """
        # Lazily try the most recent reports folder.
        candidates = []
        if tracking_file is not None:
            candidates.append(Path(tracking_file))
        else:
            workspace = Path("C:/Users/molpa/Desktop/Hermes")
            if workspace.exists():
                for p in sorted(workspace.glob("reports_*/Clonality_Tracking.xlsx"), reverse=True):
                    candidates.append(p)
        for path in candidates:
            if path.exists():
                try:
                    rows = self._excel_to_rows(path)
                except Exception:
                    rows = []
                if rows:
                    self.load_from_entries(rows)
                    return len(rows)
        # Fallback: synth
        self.set_inline_synth_entries()
        return len(self._rows)

    def set_inline_synth_entries(self) -> None:
        demo = _INLINE_SYNTH_ENTRIES()
        self.load_from_entries(demo)

    # ---- Internal helpers -----------------------------------------------

    def _refresh_btn_clicked(self) -> None:
        self.load_batch_from_tracking()

    def _refresh_table(self) -> None:
        self._table.setRowCount(0)
        rows = self._rows
        if self._disagreements_only.isChecked():
            rows = [r for r in rows if _classify_row(r, self._ml_present) != "agree"]
        self._table.setRowCount(len(rows))
        for r, entry in enumerate(rows):
            kind = _classify_row(entry, self._ml_present)
            color = _color_for(kind)
            cells = (
                str(entry.get("dit") or entry.get("DIT") or entry.get("DIT") or ""),
                str(entry.get("assay") or entry.get("Assay") or ""),
                str(entry.get("group") or entry.get("Group") or ""),
                str(entry.get("ClonalitySuggestion") or ""),
                str(entry.get("ClonalityMLSuggestion") or "(off)") if self._ml_present else "(off)",
                f"{float(entry.get('ClonalityMLConfidence') or entry.get('ClonalityConfidence') or 0.0):.2f}" if entry.get("ClonalityConfidence") else "",
                "yes" if entry.get("ClonalityReviewNeeded") else "no",
                f"{float(entry.get('DominantPeakBasepairs') or entry.get('dominant_peak_basepairs') or 0):.1f}" if entry.get("dominant_peak_basepairs") else "",
                str(entry.get("ClonalityEvidence") or entry.get("Evidence") or ""),
            )
            for c, text in enumerate(cells):
                item = QTableWidgetItem(text)
                brush = QBrush(color)
                item.setBackground(brush)
                item.setData(Qt.ItemDataRole.ToolTipRole, str(entry.get("ClonalityEvidence") or ""))
                # Ensure foreground readable on coloured background
                item.setForeground(QBrush(QColor("#0b1220")))
                self._table.setItem(r, c, item)

    def _refresh_status(self) -> None:
        if not self._rows:
            self._status_label.setText("0 rows shown — no batch loaded yet.")
            return
        total = len(self._rows)
        if self._ml_present:
            n_agree = sum(1 for e in self._rows if _classify_row(e, True) == "agree")
            n_disagree = sum(1 for e in self._rows if _classify_row(e, True) == "disagree")
            n_review = sum(1 for e in self._rows if _classify_row(e, True) == "force_review")
            self._status_label.setText(
                f"Total: {total}   |   Agree: {n_agree}   |   "
                f"Disagree (route to review): {n_disagree}   |   "
                f"Force-review: {n_review}"
            )
        else:
            self._status_label.setText(
                f"Total: {total} rows   |   ML off (interpretation.enabled=False).   Showing rule-only colors."
            )

    def _excel_to_rows(self, path) -> list[dict[str, Any]]:
        """Best-effort loader for the tracking workbook.

        Returns [] when openpyxl is unavailable; tests use
        set_inline_synth_entries() instead.
        """
        try:
            import openpyxl  # type: ignore
        except ImportError:
            return []
        try:
            book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return []
        ws = book.active
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return []
        headers = [str(c or "").strip() for c in first_row]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            entry: dict[str, Any] = dict(zip(headers, row))
            entry.setdefault("dit", entry.get("DIT"))
            entry.setdefault("assay", entry.get("Assay"))
            entry.setdefault("group", entry.get("Group"))
            rows.append(entry)
        return rows



def _INLINE_SYNTH_ENTRIES() -> list[dict[str, Any]]:
    """Eight-row demo used by `set_inline_synth_entries`."""
    base = [
        # (rule, ml, conf, review, dom_bp, evidence)
        ("polyklonal", "polyklonal", 0.92, False, "100-150 range broad humps", 110.5),
        ("monoklonal", "monoklonal", 0.95, False, "Strong single peak at FR1 expected bp", 312.4),
        ("polyklonal", "monoklonal", 0.71, False, "ML flag rare class but conf low", 110.0),
        ("monoklonal", "polyklonal", 0.83, False, "Rule mono but ML says poly", 267.1),
        ("usikker_review", "monoklonal", 0.66, True, "Ladder_qc=fail forced review", 0.0),
        ("usikker_review", "usikker_review", 0.45, True, "Kontroll_avvik on input-DNA", 0.0),
        ("bi_oligoklonal", "bi_oligoklonal", 0.78, False, "Two comparable peaks IGK", 250.8),
        ("pseudoklonal", "monoklonal", 0.74, False, "Off-window peak - rule says pseudo", 200.2),
    ]
    out: list[dict[str, Any]] = []
    for i, (rule, ml, conf, review, evidence, dom) in enumerate(base):
        out.append(
            {
                "dit": f"26OUM{i+1:05d}",
                "assay": "FR1" if i % 2 == 0 else "TCRG-A",
                "group": "(patient)" if i < 6 else "PK",
                "ClonalitySuggestion": rule,
                "ClonalityMLSuggestion": ml,
                "ClonalityConfidence": conf,
                "ClonalityMLConfidence": conf,
                "ClonalityReviewNeeded": review,
                "dominant_peak_basepairs": dom,
                "ClonalityEvidence": evidence,
            }
        )
    return out
