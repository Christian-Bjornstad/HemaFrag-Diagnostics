from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CARD_BLUE = PatternFill("solid", fgColor="D9EAF7")
CARD_TEAL = PatternFill("solid", fgColor="DDF4F1")
CARD_ORANGE = PatternFill("solid", fgColor="FCE4D6")
CARD_RED = PatternFill("solid", fgColor="FBE5E7")
CARD_GOLD = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F2")
BOX_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)
ACCEPTED_LADDER_STATUSES = {"ok", "manual_adjustment"}
GENERATED_SHEETS = {
    "Dashboard",
    "Dashboard_Data",
    "Assay_Summary",
    "Run_Summary",
    "Control_Summary",
    "PK_Sample_Delta",
    "PK_Ladder_Delta",
    "QC_Run_Trends",
    "QC_Control_Signals",
    "QC_Baseline_Config",
}
OBSOLETE_TRACKING_COLUMNS = {
    "LadderEngine",
    "LadderReasonCodes",
    "SourceFsaSha256",
    "ManualAdjustmentSha256",
    "AnalysisVersion",
    "PullUpCandidate",
    "SaturationCandidate",
}


def refresh_clonality_tracking_dashboard(
    excel_path: Path,
    *,
    dashboard_title: str = "HemaFrag Klonalitet Dashboard",
) -> None:
    if not excel_path.exists():
        return

    with pd.ExcelFile(excel_path, engine="openpyxl") as xls:
        required = {"Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"}
        if not required.issubset(set(xls.sheet_names)):
            return
        runs = pd.read_excel(excel_path, sheet_name="Runs", engine="openpyxl")
        peaks = pd.read_excel(excel_path, sheet_name="PK_Peaks", engine="openpyxl")

    work = _prepared_runs(runs)
    assay_summary = _build_assay_summary(work)
    pk_summary = _build_pk_summary(peaks)

    wb = load_workbook(excel_path)
    try:
        for name in list(wb.sheetnames):
            if name in GENERATED_SHEETS:
                del wb[name]
        for sheet_name in ("Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"):
            ws = wb[sheet_name]
            _remove_obsolete_columns(ws)
            if sheet_name != "PK_Peaks":
                _ensure_manual_adjustment_column(ws)
            _ensure_abs_delta_column(ws)
            _style_tracking_data_sheet(ws)

        dashboard = wb.create_sheet("Dashboard", 0)
        _build_dashboard(
            dashboard,
            dashboard_title=dashboard_title,
            runs=work,
            assay_summary=assay_summary,
            pk_summary=pk_summary,
        )
        _add_dashboard_charts(
            dashboard,
            assay_count=min(len(assay_summary), 12),
            pk_count=min(len(pk_summary), 12),
        )
        wb.save(excel_path)
    finally:
        wb.close()


def _text(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series("", index=frame.index, dtype="object")
    return frame[column].fillna("").astype(str).str.strip()


def _numeric(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(np.nan, index=frame.index, dtype=float)
    return pd.to_numeric(frame[column], errors="coerce")


def _prepared_runs(runs: pd.DataFrame) -> pd.DataFrame:
    work = runs.copy()
    work["_Assay"] = _text(work, "Assay").replace("", "Unknown")
    work["_SampleKind"] = _text(work, "SampleKind").str.lower()
    work["_Patient"] = work["_SampleKind"].eq("patient")
    work["_Control"] = work["_SampleKind"].eq("control") | _text(
        work,
        "Control",
    ).ne("")
    work["_LadderStatus"] = _text(work, "LadderQC").str.lower()
    work["_Manual"] = (
        work["_LadderStatus"].eq("manual_adjustment")
        | _text(work, "LadderFitStrategy").str.lower().eq("manual_adjustment")
    )
    work["_Review"] = work["_LadderStatus"].ne("") & ~work[
        "_LadderStatus"
    ].isin(ACCEPTED_LADDER_STATUSES)
    linear_r2 = _numeric(work, "LadderLinearR2")
    work["_R2"] = linear_r2.where(linear_r2.notna(), _numeric(work, "LadderR2"))
    expected = _numeric(work, "LadderExpectedStepCount")
    fitted = _numeric(work, "LadderFittedStepCount")
    work["_Partial"] = expected.gt(0) & fitted.lt(expected)
    return work


def _build_assay_summary(work: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Assay",
        "Files",
        "PatientFiles",
        "ControlFiles",
        "ReviewFiles",
        "ManualAdjusted",
        "AvgR2",
        "PartialFits",
        "ReviewRate",
    ]
    rows = []
    for assay, group in work.groupby("_Assay", sort=True):
        rows.append(
            {
                "Assay": assay,
                "Files": len(group),
                "PatientFiles": int(group["_Patient"].sum()),
                "ControlFiles": int(group["_Control"].sum()),
                "ReviewFiles": int(group["_Review"].sum()),
                "ManualAdjusted": int(group["_Manual"].sum()),
                "AvgR2": group["_R2"].mean(),
                "PartialFits": int(group["_Partial"].sum()),
                "ReviewRate": float(group["_Review"].mean()),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _build_pk_summary(peaks: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Assay",
        "MarkerRows",
        "MeanAbsDeltaBP",
        "MaxAbsDeltaBP",
        "Over2bp",
        "AvgHeight",
    ]
    if peaks.empty:
        return pd.DataFrame(columns=columns)
    work = peaks.copy()
    kind = _text(work, "Kind").str.lower()
    assay = _text(work, "Assay")
    work = work.loc[kind.eq("sample") & assay.ne("SL")].copy()
    work["_Assay"] = _text(work, "Assay").replace("", "Unknown")
    delta = _numeric(work, "AbsDeltaBP")
    if delta.isna().all():
        delta = _numeric(work, "DeltaBP").abs()
    work["_AbsDelta"] = delta
    work["_Height"] = _numeric(work, "Height")
    rows = []
    for assay_name, group in work.groupby("_Assay", sort=True):
        rows.append(
            {
                "Assay": assay_name,
                "MarkerRows": len(group),
                "MeanAbsDeltaBP": group["_AbsDelta"].mean(),
                "MaxAbsDeltaBP": group["_AbsDelta"].max(),
                "Over2bp": int(group["_AbsDelta"].gt(2).sum()),
                "AvgHeight": group["_Height"].mean(),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _scope_row(label: str, frame: pd.DataFrame) -> dict:
    return {
        "Scope": label,
        "Runs": len(frame),
        "Accepted Fits": int(
            frame["_LadderStatus"].isin(ACCEPTED_LADDER_STATUSES).sum()
        ),
        "Review Required": int(frame["_Review"].sum()),
        "Manual Adjusted": int(frame["_Manual"].sum()),
        "Accepted Rate": (
            float(frame["_LadderStatus"].isin(ACCEPTED_LADDER_STATUSES).mean())
            if not frame.empty
            else 0.0
        ),
        "Avg R2": frame["_R2"].mean(),
        "Partial Fits": int(frame["_Partial"].sum()),
    }


def _build_dashboard(
    ws,
    *,
    dashboard_title: str,
    runs: pd.DataFrame,
    assay_summary: pd.DataFrame,
    pk_summary: pd.DataFrame,
) -> None:
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, 18):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["I"].width = 14

    ws.merge_cells("A1:Q2")
    ws["A1"] = dashboard_title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    latest_date = _text(runs, "RunDate").max() if not runs.empty else ""
    metrics = [
        ("Tracked files", len(runs), CARD_BLUE),
        ("Unique assays", int(runs["_Assay"].nunique()), CARD_TEAL),
        ("Review files", int(runs["_Review"].sum()), CARD_RED),
        ("Manual adjusted", int(runs["_Manual"].sum()), CARD_GOLD),
        (
            "Accepted ladder rate",
            float(runs["_LadderStatus"].isin(ACCEPTED_LADDER_STATUSES).mean())
            if not runs.empty
            else 0.0,
            CARD_TEAL,
        ),
        ("PK marker rows", int(len(pk_summary)), CARD_BLUE),
        ("Partial fits", int(runs["_Partial"].sum()), CARD_ORANGE),
        ("Latest run", latest_date, CARD_GOLD),
    ]
    for index, (label, value, fill) in enumerate(metrics, start=1):
        ws.cell(4, index, label)
        ws.cell(5, index, value)
        for row in (4, 5):
            cell = ws.cell(row, index)
            cell.fill = fill
            cell.border = BOX_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(4, index).font = BOLD
    ws["E5"].number_format = "0.0%"

    ws["A7"] = "Ladder Overview"
    ws["A7"].font = Font(size=13, bold=True)
    overview = pd.DataFrame(
        [
            _scope_row("All", runs),
            _scope_row("Patient", runs.loc[runs["_Patient"]]),
            _scope_row("Control", runs.loc[runs["_Control"]]),
        ]
    )
    _write_dashboard_frame(ws, overview, start_row=8, start_col=1)

    ws["J7"] = "PK Sample Delta Focus"
    ws["J7"].font = Font(size=13, bold=True)
    _write_dashboard_frame(
        ws,
        pk_summary.head(12),
        start_row=8,
        start_col=10,
    )

    ws["A15"] = "Assay Watchlist"
    ws["A15"].font = Font(size=13, bold=True)
    _write_dashboard_frame(
        ws,
        assay_summary.head(20),
        start_row=16,
        start_col=1,
    )
    watch_end = 16 + max(min(len(assay_summary), 20), 1)
    ws.conditional_formatting.add(
        f"E17:E{watch_end}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="E2F0D9",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFE699",
            end_type="max",
            end_color="F4CCCC",
        ),
    )
    ws["F9"].number_format = "0.0%"
    ws["F10"].number_format = "0.0%"
    ws["F11"].number_format = "0.0%"
    for row_idx in range(17, watch_end + 1):
        ws.cell(row_idx, 9).number_format = "0.0%"
    ws.freeze_panes = "A8"


def _write_dashboard_frame(
    ws,
    frame: pd.DataFrame,
    *,
    start_row: int,
    start_col: int,
) -> None:
    for offset, header in enumerate(frame.columns):
        ws.cell(start_row, start_col + offset, header)
    for row_offset, row in enumerate(
        frame.itertuples(index=False, name=None),
        start=1,
    ):
        for col_offset, value in enumerate(row):
            ws.cell(
                start_row + row_offset,
                start_col + col_offset,
                None if pd.isna(value) else value,
            )
    _style_table(
        ws,
        start_row,
        start_row + max(len(frame), 1),
        start_col,
        start_col + max(len(frame.columns), 1) - 1,
    )


def _add_dashboard_charts(ws, *, assay_count: int, pk_count: int) -> None:
    if assay_count:
        assay_chart = BarChart()
        assay_chart.title = "Files by Assay"
        assay_chart.y_axis.title = "Files"
        assay_chart.height = 7
        assay_chart.width = 11
        assay_chart.add_data(
            Reference(
                ws,
                min_col=2,
                min_row=16,
                max_row=16 + assay_count,
            ),
            titles_from_data=True,
        )
        assay_chart.set_categories(
            Reference(ws, min_col=1, min_row=17, max_row=16 + assay_count)
        )
        assay_chart.legend = None
        assay_chart.series[0].tx = SeriesLabel(v="Files")
        ws.add_chart(assay_chart, "A35")

    if pk_count:
        pk_chart = BarChart()
        pk_chart.title = "PK sample mean |delta bp|"
        pk_chart.y_axis.title = "|delta bp|"
        pk_chart.height = 7
        pk_chart.width = 11
        pk_chart.add_data(
            Reference(
                ws,
                min_col=12,
                min_row=8,
                max_row=8 + pk_count,
            ),
            titles_from_data=True,
        )
        pk_chart.set_categories(
            Reference(ws, min_col=10, min_row=9, max_row=8 + pk_count)
        )
        pk_chart.legend = None
        pk_chart.series[0].tx = SeriesLabel(v="Mean |delta bp|")
        ws.add_chart(pk_chart, "J35")


def _remove_obsolete_columns(ws) -> None:
    for column_index in range(ws.max_column, 0, -1):
        header = str(ws.cell(1, column_index).value or "").strip()
        if header in OBSOLETE_TRACKING_COLUMNS:
            ws.delete_cols(column_index)


def _ensure_manual_adjustment_column(ws) -> None:
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    if "ManualAdjustmentUsed" in headers:
        return
    strategy_col = headers.get("LadderFitStrategy")
    if not strategy_col:
        return
    target_col = strategy_col + 1
    ws.insert_cols(target_col)
    ws.cell(1, target_col, "ManualAdjustmentUsed")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(
            row_idx,
            target_col,
            str(ws.cell(row_idx, strategy_col).value or "").strip().lower()
            == "manual_adjustment",
        )


def _ensure_abs_delta_column(ws) -> None:
    if ws.title != "PK_Peaks":
        return
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    if "AbsDeltaBP" in headers:
        return
    delta_col = headers.get("DeltaBP")
    if not delta_col:
        return
    target_col = ws.max_column + 1
    ws.cell(1, target_col, "AbsDeltaBP")
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row_idx, delta_col).value
        try:
            ws.cell(row_idx, target_col, abs(float(value)))
        except (TypeError, ValueError):
            ws.cell(row_idx, target_col, None)


def _style_table(
    ws,
    header_row: int,
    data_end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(header_row, col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX_BORDER
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=max(data_end_row, header_row + 1),
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = BOX_BORDER
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFB")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def _style_tracking_data_sheet(ws) -> None:
    if ws.max_column < 1:
        return
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = BOX_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "G2" if ws.title != "PK_Peaks" else "F2"
    ws.auto_filter.ref = ws.dimensions

    identity_col = headers.get("IdentityKey")
    if identity_col:
        ws.column_dimensions[get_column_letter(identity_col)].hidden = True

    sample_end = min(ws.max_row, 200)
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        width = len(header) + 2
        for row_idx in range(2, sample_end + 1):
            width = max(width, len(str(ws.cell(row_idx, col_idx).value or "")) + 1)
        cap = 42 if header in {"File", "SourceRunDir", "Reason"} else 24
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width, 10),
            cap,
        )
        normalized = header.lower()
        if "r2" in normalized:
            number_format = "0.000000"
        elif "bp" in normalized or "area" in normalized or "height" in normalized:
            number_format = "0.00"
        else:
            number_format = None
        if number_format:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = number_format

    status_col = headers.get("LadderQC")
    if status_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, status_col)
            status = str(cell.value or "").strip().lower()
            if status in ACCEPTED_LADDER_STATUSES:
                cell.fill = CARD_TEAL
            elif status:
                cell.fill = CARD_RED
