"""Build and merge deterministic chemist-labeling batches."""
from __future__ import annotations

import hashlib
import json
import os
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from core.analyses.clonality.ml_data_contract import (
    CHEMIST_LABEL_COLUMN,
    load_tracking_run_table,
)
from core.analyses.clonality.ml_training import (
    ANNOTATION_CLASSES_ORDER,
    normalize_annotation_label,
)


LABELING_BATCH_SCHEMA_VERSION = "clonality_labeling_batch_v1"
LABELING_BATCH_COLUMNS = (
    "LabelingBatchId",
    "LabelingBatchRank",
    "LabelingAssayRank",
    "LabelingSelectionReason",
    "LabelingFeatureDistance",
    "LabelingRuleStratum",
    "LabelingSourceRunKey",
)
DEFAULT_EXCLUDED_ASSAYS = {"IKZF1", "Ktr-albumin"}

_SELECTION_NON_FEATURE_COLUMNS = {
    "RuleConfidence",
    "RuleReviewNeeded",
    "ClonalityConfidence",
    "ClonalityReviewNeeded",
}


@dataclass(frozen=True)
class LabelingBatch:
    rows: pd.DataFrame
    assay_summary: pd.DataFrame
    rule_summary: pd.DataFrame
    manifest: dict[str, Any]


def build_clonality_labeling_batch(
    tracking_rows: pd.DataFrame,
    feature_rows: pd.DataFrame,
    *,
    batch_id: str,
    per_assay: int = 24,
    max_rows: int | None = None,
    review_fraction: float = 0.65,
    random_state: int = 20260726,
    exclude_assays: set[str] | None = None,
) -> LabelingBatch:
    """Select a deterministic, diverse, unlabeled cohort for chemist review."""
    if per_assay < 1:
        raise ValueError("per_assay must be at least 1")
    if max_rows is not None and max_rows < 1:
        raise ValueError("max_rows must be at least 1")
    if not 0.0 <= review_fraction <= 1.0:
        raise ValueError("review_fraction must be between 0 and 1")
    if not str(batch_id or "").strip():
        raise ValueError("batch_id must not be empty")

    tracking = tracking_rows.copy()
    features = feature_rows.copy()
    required_tracking = {"IdentityKey", "DIT", "Assay", "SourceRunDir"}
    required_features = {"IdentityKey", "Assay", "SourceRunKey"}
    _require_columns(tracking, required_tracking, "tracking rows")
    _require_columns(features, required_features, "feature rows")

    if CHEMIST_LABEL_COLUMN not in tracking.columns:
        tracking[CHEMIST_LABEL_COLUMN] = ""
    excluded = DEFAULT_EXCLUDED_ASSAYS if exclude_assays is None else set(exclude_assays)
    if excluded and "Assay" in tracking.columns:
        assay_names = tracking["Assay"].fillna("").astype(str).str.strip()
        tracking = tracking.loc[~assay_names.isin(excluded)].copy()
    labels = (
        tracking[CHEMIST_LABEL_COLUMN]
        .fillna("")
        .map(normalize_annotation_label)
    )
    tracking = tracking.loc[labels.eq("")].copy()
    if tracking.empty:
        raise ValueError("tracking workbook has no unlabeled patient rows")

    tracking["_AssayKey"] = tracking["Assay"].map(_assay_key)
    features["_AssayKey"] = features["Assay"].map(_assay_key)
    join_columns = ["IdentityKey", "_AssayKey"]
    _reject_duplicate_keys(tracking, join_columns, "tracking")
    _reject_duplicate_keys(features, join_columns, "features")

    numeric_feature_columns = [
        column
        for column in features.columns
        if column not in _SELECTION_NON_FEATURE_COLUMNS
        and column not in {
            "IdentityKey",
            "DIT",
            "Assay",
            "SourceRunKey",
            "FsaSourceHash",
            "FsaContentHash",
            CHEMIST_LABEL_COLUMN,
            "RuleSuggestion",
            "_AssayKey",
        }
        and (
            pd.api.types.is_numeric_dtype(features[column])
            or pd.api.types.is_bool_dtype(features[column])
        )
        and pd.to_numeric(features[column], errors="coerce").nunique(dropna=True) > 1
    ]
    if not numeric_feature_columns:
        raise ValueError("feature artifact has no varying numeric features")

    feature_context_columns = [
        *join_columns,
        "SourceRunKey",
        "RuleSuggestion",
        "RuleConfidence",
        "RuleReviewNeeded",
        *numeric_feature_columns,
    ]
    feature_context_columns = [
        column for column in feature_context_columns if column in features.columns
    ]
    merged = tracking.merge(
        features[feature_context_columns],
        on=join_columns,
        how="left",
        validate="one_to_one",
        indicator=True,
        suffixes=("", "_Feature"),
    )
    unmatched = merged["_merge"].ne("both")
    if unmatched.any():
        raise ValueError(
            f"{int(unmatched.sum())} unlabeled tracking row(s) have no feature row"
        )
    merged = merged.drop(columns=["_merge"])
    merged["_RuleSuggestion"] = _first_text_column(
        merged,
        "RuleSuggestion",
        "ClonalitySuggestion",
    )
    merged["_RuleConfidence"] = _first_numeric_column(
        merged,
        "RuleConfidence",
        "ClonalityConfidence",
    )
    merged["_RuleReviewNeeded"] = _first_bool_column(
        merged,
        "RuleReviewNeeded",
        "ClonalityReviewNeeded",
    )
    merged["_SourceRunKey"] = (
        merged.get("SourceRunKey", pd.Series("", index=merged.index))
        .fillna("")
        .astype(str)
        .str.strip()
    )

    capacities = {
        assay: min(per_assay, int(len(group)))
        for assay, group in merged.groupby("_AssayKey", sort=True)
        if assay
    }
    quotas = _balanced_quotas(capacities, max_rows=max_rows)
    selected_by_assay: dict[str, pd.DataFrame] = {}
    for assay, quota in quotas.items():
        if quota <= 0:
            continue
        group = merged.loc[merged["_AssayKey"].eq(assay)].copy()
        selected_by_assay[assay] = _select_diverse_rows(
            group,
            numeric_feature_columns,
            quota=quota,
            review_fraction=review_fraction,
            random_state=random_state,
        )

    ordered_pieces: list[pd.DataFrame] = []
    max_assay_rows = max((len(frame) for frame in selected_by_assay.values()), default=0)
    global_rank = 0
    for assay_rank in range(max_assay_rows):
        for assay in sorted(selected_by_assay):
            frame = selected_by_assay[assay]
            if assay_rank >= len(frame):
                continue
            global_rank += 1
            row = frame.iloc[[assay_rank]].copy()
            row["LabelingBatchRank"] = global_rank
            ordered_pieces.append(row)
    if not ordered_pieces:
        raise ValueError("labeling selection produced no rows")
    selected = pd.concat(ordered_pieces, ignore_index=True, sort=False)
    selected = _expand_selected_to_parallel_rows(selected, merged)

    selected["LabelingBatchId"] = str(batch_id).strip()
    selected["LabelingRuleStratum"] = selected["_RuleSuggestion"]
    selected["LabelingSourceRunKey"] = selected["_SourceRunKey"]
    selected["ClonalitySuggestion"] = selected["_RuleSuggestion"]
    selected["ClonalityConfidence"] = selected["_RuleConfidence"]
    selected["ClonalityReviewNeeded"] = selected["_RuleReviewNeeded"]
    selected[CHEMIST_LABEL_COLUMN] = ""

    original_columns = [
        column
        for column in tracking_rows.columns
        if not str(column).startswith("_")
    ]
    output_columns = list(
        dict.fromkeys(
            [
                *original_columns,
                "ClonalitySuggestion",
                "ClonalityConfidence",
                "ClonalityReviewNeeded",
                *LABELING_BATCH_COLUMNS,
            ]
        )
    )
    rows = selected.reindex(columns=output_columns).copy()
    rows = rows.sort_values("LabelingBatchRank", kind="stable").reset_index(drop=True)

    assay_summary = (
        rows.groupby("Assay", dropna=False, sort=True)
        .agg(
            SelectedRows=("IdentityKey", "size"),
            DistinctDITs=("DIT", "nunique"),
            SourceRuns=("LabelingSourceRunKey", "nunique"),
            RuleReviewRows=("ClonalityReviewNeeded", "sum"),
            RuleStrata=("LabelingRuleStratum", "nunique"),
        )
        .reset_index()
    )
    assay_summary["RuleReviewFraction"] = (
        assay_summary["RuleReviewRows"] / assay_summary["SelectedRows"]
    )
    rule_summary = (
        rows.groupby(
            ["Assay", "LabelingRuleStratum", "ClonalityReviewNeeded"],
            dropna=False,
            sort=True,
        )
        .agg(
            SelectedRows=("IdentityKey", "size"),
            DistinctDITs=("DIT", "nunique"),
            SourceRuns=("LabelingSourceRunKey", "nunique"),
        )
        .reset_index()
    )
    manifest = {
        "schema_version": LABELING_BATCH_SCHEMA_VERSION,
        "batch_id": str(batch_id).strip(),
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "random_state": int(random_state),
        "per_assay": int(per_assay),
        "max_rows": int(max_rows) if max_rows is not None else None,
        "review_fraction_target": float(review_fraction),
        "eligible_unlabeled_rows": int(len(merged)),
        "selected_rows": int(len(rows)),
        "selected_assays": int(rows["Assay"].nunique()),
        "selected_dits": int(rows["DIT"].nunique()),
        "selected_source_runs": int(rows["LabelingSourceRunKey"].nunique()),
        "selected_rule_review_rows": int(
            rows["ClonalityReviewNeeded"].fillna(False).astype(bool).sum()
        ),
        "selected_rule_strata": int(rows["LabelingRuleStratum"].nunique()),
        "selection_feature_count": int(len(numeric_feature_columns)),
        "rule_suggestions_used_as_labels": False,
        "excluded_assays": sorted(excluded),
    }
    return LabelingBatch(
        rows=rows,
        assay_summary=assay_summary,
        rule_summary=rule_summary,
        manifest=manifest,
    )


def write_clonality_labeling_batch(
    batch: LabelingBatch,
    output_path: Path | str,
    *,
    source_workbook: Path | str,
    source_features: Path | str,
    overwrite: bool = False,
) -> dict[str, Path]:
    """Write a GUI-compatible batch workbook and local provenance manifest."""
    output = Path(output_path).expanduser()
    if output.exists() and not overwrite:
        raise FileExistsError(f"output workbook already exists: {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    metadata = {
        **batch.manifest,
        "source_workbook": str(Path(source_workbook).expanduser().resolve()),
        "source_features": str(Path(source_features).expanduser().resolve()),
        "output_workbook": str(output.resolve()),
    }
    metadata_frame = pd.DataFrame(
        [{"Key": key, "Value": _json_cell(value)} for key, value in metadata.items()]
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        batch.rows.to_excel(writer, sheet_name="Runs", index=False)
        batch.rows.to_excel(writer, sheet_name="Patient_Runs", index=False)
        batch.assay_summary.to_excel(writer, sheet_name="Batch_Summary", index=False)
        batch.rule_summary.to_excel(writer, sheet_name="Rule_Summary", index=False)
        metadata_frame.to_excel(writer, sheet_name="Batch_Metadata", index=False)
    _format_batch_workbook(output)

    manifest_path = output.with_suffix(".manifest.json")
    _atomic_write_text(
        manifest_path,
        json.dumps(metadata, indent=2, sort_keys=True, ensure_ascii=True) + "\n",
    )
    return {"workbook": output, "manifest": manifest_path}


def merge_clonality_labeling_batch(
    batch_workbook: Path | str,
    target_workbook: Path | str,
    *,
    allow_overwrite: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Merge reviewed batch labels into the full workbook with conflict guards."""
    batch_path = Path(batch_workbook).expanduser()
    target_path = Path(target_workbook).expanduser()
    batch = load_tracking_run_table(batch_path).frame
    target = load_tracking_run_table(target_path).frame
    _require_columns(
        batch,
        {"IdentityKey", "Assay", CHEMIST_LABEL_COLUMN},
        "batch workbook",
    )
    _require_columns(target, {"IdentityKey", "Assay"}, "target workbook")
    if CHEMIST_LABEL_COLUMN not in target.columns:
        target[CHEMIST_LABEL_COLUMN] = ""

    batch = batch.copy()
    target = target.copy()
    batch["_AssayKey"] = batch["Assay"].map(_assay_key)
    target["_AssayKey"] = target["Assay"].map(_assay_key)
    key_columns = ["IdentityKey", "_AssayKey"]
    labels = (
        batch[CHEMIST_LABEL_COLUMN]
        .fillna("")
        .map(normalize_annotation_label)
    )
    labeled = batch.loc[labels.ne(""), key_columns + [CHEMIST_LABEL_COLUMN]].copy()
    labeled[CHEMIST_LABEL_COLUMN] = labels.loc[labels.ne("")]
    invalid = sorted(
        set(labeled[CHEMIST_LABEL_COLUMN]) - set(ANNOTATION_CLASSES_ORDER)
    )
    if invalid:
        raise ValueError(f"batch contains invalid chemist labels: {', '.join(invalid)}")
    duplicate = labeled.duplicated(subset=key_columns, keep=False)
    if duplicate.any():
        conflicts = (
            labeled.loc[duplicate]
            .groupby(key_columns)[CHEMIST_LABEL_COLUMN]
            .nunique()
        )
        if conflicts.gt(1).any():
            raise ValueError("batch contains conflicting labels for one IdentityKey+Assay")
        labeled = labeled.drop_duplicates(subset=key_columns, keep="last")

    target_key_map = {
        (str(row["IdentityKey"]), str(row["_AssayKey"])): _clean_text(
            row.get(CHEMIST_LABEL_COLUMN)
        )
        for _, row in target.iterrows()
    }
    updates: dict[tuple[str, str], str] = {}
    unchanged = 0
    conflicts: list[dict[str, str]] = []
    missing: list[dict[str, str]] = []
    for _, row in labeled.iterrows():
        key = (str(row["IdentityKey"]), str(row["_AssayKey"]))
        new_label = str(row[CHEMIST_LABEL_COLUMN]).strip()
        if key not in target_key_map:
            missing.append({"IdentityKey": key[0], "AssayKey": key[1]})
            continue
        current_label = target_key_map[key]
        if current_label == new_label:
            unchanged += 1
        elif current_label and not allow_overwrite:
            conflicts.append(
                {
                    "IdentityKey": key[0],
                    "AssayKey": key[1],
                    "TargetLabel": current_label,
                    "BatchLabel": new_label,
                }
            )
        else:
            updates[key] = new_label

    if updates and not dry_run:
        _write_label_updates(target_path, updates)
    return {
        "batch_labeled_rows": int(len(labeled)),
        "labels_written": int(len(updates)),
        "labels_unchanged": int(unchanged),
        "conflict_count": int(len(conflicts)),
        "missing_target_count": int(len(missing)),
        "allow_overwrite": bool(allow_overwrite),
        "dry_run": bool(dry_run),
        "conflicts": conflicts,
        "missing": missing,
    }


def _select_diverse_rows(
    group: pd.DataFrame,
    feature_columns: list[str],
    *,
    quota: int,
    review_fraction: float,
    random_state: int,
) -> pd.DataFrame:
    working = group.copy().reset_index(drop=True)
    matrix = _standardized_matrix(working, feature_columns)
    stable_ties = np.asarray(
        [
            _stable_unit_interval(random_state, value)
            for value in working["IdentityKey"].astype(str)
        ],
        dtype=float,
    )
    review = working["_RuleReviewNeeded"].fillna(False).astype(bool).to_numpy()
    confidence = (
        pd.to_numeric(working["_RuleConfidence"], errors="coerce")
        .fillna(0.0)
        .clip(0.0, 1.0)
        .to_numpy(dtype=float)
    )
    rule = working["_RuleSuggestion"].fillna("").astype(str).to_numpy()
    source_run = working["_SourceRunKey"].fillna("").astype(str).to_numpy()

    selected: list[int] = []
    selected_rules: set[str] = set()
    selected_runs: set[str] = set()
    reasons: dict[int, str] = {}
    distances: dict[int, float] = {}
    while len(selected) < min(quota, len(working)):
        remaining = np.asarray(
            [index for index in range(len(working)) if index not in selected],
            dtype=int,
        )
        if selected:
            deltas = matrix[remaining, None, :] - matrix[np.asarray(selected)][None, :, :]
            min_distance = np.sqrt(np.mean(np.square(deltas), axis=2)).min(axis=1)
            max_distance = float(np.max(min_distance)) if min_distance.size else 0.0
            diversity = min_distance / max_distance if max_distance > 0 else min_distance
        else:
            min_distance = np.zeros(len(remaining), dtype=float)
            diversity = np.zeros(len(remaining), dtype=float)

        desired_review_count = int(np.ceil((len(selected) + 1) * review_fraction))
        need_review = sum(bool(review[index]) for index in selected) < desired_review_count
        unseen_rule = np.asarray(
            [bool(rule[index]) and rule[index] not in selected_rules for index in remaining],
            dtype=float,
        )
        unseen_run = np.asarray(
            [
                bool(source_run[index]) and source_run[index] not in selected_runs
                for index in remaining
            ],
            dtype=float,
        )
        review_priority = np.asarray(
            [float(need_review and review[index]) for index in remaining],
            dtype=float,
        )
        low_confidence = 1.0 - confidence[remaining]
        score = (
            2.0 * diversity
            + 1.2 * unseen_rule
            + 0.8 * unseen_run
            + 0.75 * review_priority
            + 0.25 * low_confidence
            + 1e-6 * stable_ties[remaining]
        )
        chosen_position = int(np.argmax(score))
        chosen = int(remaining[chosen_position])
        selected.append(chosen)
        if rule[chosen]:
            selected_rules.add(rule[chosen])
        if source_run[chosen]:
            selected_runs.add(source_run[chosen])
        reason_parts = []
        if not reasons:
            reason_parts.append("review_low_confidence_seed")
        else:
            reason_parts.append("feature_diversity")
        if unseen_rule[chosen_position]:
            reason_parts.append("new_rule_stratum")
        if unseen_run[chosen_position]:
            reason_parts.append("new_source_run")
        if review_priority[chosen_position]:
            reason_parts.append("rule_review_priority")
        reasons[chosen] = "+".join(reason_parts)
        distances[chosen] = float(min_distance[chosen_position])

    result = working.iloc[selected].copy().reset_index(drop=True)
    result["LabelingAssayRank"] = np.arange(1, len(result) + 1, dtype=int)
    result["LabelingSelectionReason"] = [reasons[index] for index in selected]
    result["LabelingFeatureDistance"] = [distances[index] for index in selected]
    return result


def _expand_selected_to_parallel_rows(
    selected: pd.DataFrame,
    eligible_rows: pd.DataFrame,
) -> pd.DataFrame:
    """Include every unlabeled row for each selected DIT+assay pair."""
    if selected.empty:
        return selected
    seed_keys = selected[["DIT", "_AssayKey"]].drop_duplicates()
    expanded = eligible_rows.merge(
        seed_keys,
        on=["DIT", "_AssayKey"],
        how="inner",
        validate="many_to_one",
    )
    seed_meta = selected[
        [
            "DIT",
            "_AssayKey",
            "LabelingBatchRank",
            "LabelingAssayRank",
            "LabelingSelectionReason",
            "LabelingFeatureDistance",
        ]
    ].drop_duplicates(subset=["DIT", "_AssayKey"], keep="first")
    expanded = expanded.merge(
        seed_meta,
        on=["DIT", "_AssayKey"],
        how="left",
        validate="many_to_one",
    )
    expanded["_ParallelOrder"] = (
        expanded.get("Well", pd.Series("", index=expanded.index))
        .fillna("")
        .astype(str)
        .str.strip()
        + "|"
        + expanded.get("File", pd.Series("", index=expanded.index))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    expanded["LabelingSelectionReason"] = expanded["LabelingSelectionReason"].where(
        expanded["IdentityKey"].isin(set(selected["IdentityKey"])),
        expanded["LabelingSelectionReason"].fillna("") + "+parallel_pair",
    )
    expanded = expanded.sort_values(
        ["LabelingBatchRank", "_ParallelOrder"],
        kind="stable",
    ).reset_index(drop=True)
    expanded["LabelingBatchRank"] = np.arange(1, len(expanded) + 1, dtype=int)
    expanded = expanded.drop(columns=["_ParallelOrder"])
    return expanded


def _standardized_matrix(frame: pd.DataFrame, columns: list[str]) -> np.ndarray:
    numeric = frame[columns].apply(pd.to_numeric, errors="coerce")
    median = numeric.median(axis=0, skipna=True).fillna(0.0)
    filled = numeric.fillna(median)
    q25 = filled.quantile(0.25)
    q75 = filled.quantile(0.75)
    scale = (q75 - q25).replace(0.0, np.nan)
    fallback = filled.std(axis=0, ddof=0).replace(0.0, np.nan)
    scale = scale.fillna(fallback).fillna(1.0)
    standardized = (filled - median) / scale
    return np.clip(
        standardized.to_numpy(dtype=float),
        -10.0,
        10.0,
    )


def _balanced_quotas(
    capacities: Mapping[str, int],
    *,
    max_rows: int | None,
) -> dict[str, int]:
    quotas = {assay: 0 for assay in sorted(capacities)}
    target = sum(int(value) for value in capacities.values())
    if max_rows is not None:
        target = min(target, int(max_rows))
    while sum(quotas.values()) < target:
        progressed = False
        for assay in sorted(quotas):
            if quotas[assay] >= int(capacities[assay]):
                continue
            quotas[assay] += 1
            progressed = True
            if sum(quotas.values()) >= target:
                break
        if not progressed:
            break
    return quotas


def _write_label_updates(
    workbook_path: Path,
    updates: Mapping[tuple[str, str], str],
) -> None:
    workbook = load_workbook(workbook_path)
    try:
        for sheet_name in ("Runs", "Run", "Patient_Runs"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = {
                str(cell.value or "").strip(): cell.column
                for cell in sheet[1]
                if str(cell.value or "").strip()
            }
            identity_column = headers.get("IdentityKey")
            assay_column = headers.get("Assay")
            if identity_column is None or assay_column is None:
                continue
            label_column = headers.get(CHEMIST_LABEL_COLUMN)
            if label_column is None:
                label_column = sheet.max_column + 1
                sheet.cell(1, label_column, CHEMIST_LABEL_COLUMN)
            for row_number in range(2, sheet.max_row + 1):
                key = (
                    str(sheet.cell(row_number, identity_column).value or "").strip(),
                    _assay_key(sheet.cell(row_number, assay_column).value),
                )
                if key in updates:
                    sheet.cell(row_number, label_column, updates[key])
        temporary = workbook_path.with_suffix(workbook_path.suffix + ".tmp")
        workbook.save(temporary)
        os.replace(temporary, workbook_path)
    finally:
        workbook.close()


def _format_batch_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font
            for column_cells in sheet.iter_cols():
                values = [str(cell.value or "") for cell in column_cells[:100]]
                width = min(max(max((len(value) for value in values), default=0) + 2, 10), 36)
                sheet.column_dimensions[column_cells[0].column_letter].width = width
        workbook.save(path)
    finally:
        workbook.close()


def _first_text_column(frame: pd.DataFrame, *columns: str) -> pd.Series:
    result = pd.Series("", index=frame.index, dtype=object)
    for column in columns:
        if column not in frame.columns:
            continue
        values = frame[column].fillna("").astype(str).str.strip()
        result = result.where(result.ne(""), values)
    return result


def _first_numeric_column(frame: pd.DataFrame, *columns: str) -> pd.Series:
    result = pd.Series(np.nan, index=frame.index, dtype=float)
    for column in columns:
        if column in frame.columns:
            result = result.fillna(pd.to_numeric(frame[column], errors="coerce"))
    return result.fillna(0.0)


def _first_bool_column(frame: pd.DataFrame, *columns: str) -> pd.Series:
    for column in columns:
        if column in frame.columns:
            return frame[column].map(_as_bool)
    return pd.Series(False, index=frame.index, dtype=bool)


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _reject_duplicate_keys(
    frame: pd.DataFrame,
    columns: list[str],
    name: str,
) -> None:
    duplicate = frame.duplicated(subset=columns, keep=False)
    if duplicate.any():
        raise ValueError(
            f"{int(duplicate.sum())} {name} row(s) duplicate {'+'.join(columns)}"
        )


def _require_columns(frame: pd.DataFrame, required: set[str], name: str) -> None:
    missing = sorted(required - set(frame.columns))
    if missing:
        raise KeyError(f"{name} missing required columns: {', '.join(missing)}")


def _assay_key(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .replace(" ", "")
        .replace("-", "")
        .replace("_", "")
        .upper()
    )


def _stable_unit_interval(random_state: int, identity: str) -> float:
    payload = f"{int(random_state)}::{identity}".encode("utf-8")
    digest = hashlib.sha256(payload).digest()
    return int.from_bytes(digest[:8], "big") / float(2**64 - 1)


def _json_cell(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, sort_keys=True, ensure_ascii=True)


def _atomic_write_text(path: Path, text: str) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(text, encoding="utf-8")
    os.replace(temporary, path)


__all__ = [
    "LABELING_BATCH_COLUMNS",
    "LABELING_BATCH_SCHEMA_VERSION",
    "LabelingBatch",
    "build_clonality_labeling_batch",
    "merge_clonality_labeling_batch",
    "write_clonality_labeling_batch",
]
