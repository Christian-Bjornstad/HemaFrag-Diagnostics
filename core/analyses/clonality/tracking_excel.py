from __future__ import annotations

import hashlib
import hmac
import os
import secrets
from pathlib import Path

import numpy as np
import pandas as pd

from config import APP_SETTINGS
from fraggler.fraggler import print_green
from core.analyses.clonality.tracking_dashboard import refresh_clonality_tracking_dashboard
from core.qc.qc_markers import (
    control_id_from_filename,
    find_peak_near_bp,
    find_peak_near_bp_with_fallback,
    make_run_key,
    markers_for_entry,
    parse_batch_from_filename,
    parse_pcr_date_from_filename,
    parse_run_code_from_filename,
    parse_well_from_filename,
)
from core.qc.qc_rules import QCRules
from core.utils import strip_stage_prefix

import threading

CLONALITY_TRACKING_FILENAME = "Clonality_Tracking.xlsx"
_clonality_excel_lock = threading.Lock()
CONTROL_IDS = {"PK", "PK1", "PK2", "NK", "RK"}
TRACKING_IDENTITY_SALT_ENV = "FRAGGLER_TRACKING_IDENTITY_SALT"
TRACKING_IDENTITY_SALT_PATH_ENV = "FRAGGLER_TRACKING_IDENTITY_SALT_PATH"
DEFAULT_TRACKING_IDENTITY_SALT_PATH = Path.home() / ".config" / "fraggler" / "tracking_identity_salt.txt"
RUN_SHEET_COLUMNS = [
    "Month",
    "IdentityKey",
    "File",
    "SourceRunDir",
    "DIT",
    "Assay",
    "SampleKind",
    "Group",
    "Control",
    "RunDate",
    "RunCode",
    "Well",
    "Batch",
    "Ladder",
    "LadderQC",
    "LadderFitStrategy",
    "LadderExpectedStepCount",
    "LadderFittedStepCount",
    "LadderR2",
    "LadderLinearR2",
    "LadderLinearMeanResidualBp",
    "LadderLinearMaxResidualBp",
    "LadderCurvature",
]
PEAK_SHEET_COLUMNS = [
    "Month",
    "IdentityKey",
    "File",
    "SourceRunDir",
    "DIT",
    "Assay",
    "Control",
    "RunDate",
    "RunCode",
    "Well",
    "Batch",
    "MarkerName",
    "Kind",
    "Channel",
    "ExpectedBP",
    "WindowBP",
    "SearchMode",
    "SearchWindowBP",
    "FoundBP",
    "DeltaBP",
    "Height",
    "Area",
    "OK",
    "Reason",
    "AbsDeltaBP",
]


def build_clonality_qc_rules() -> QCRules:
    qc_settings = APP_SETTINGS.get("qc", {})
    sample_window = qc_settings.get("sample_peak_window_bp", qc_settings.get("w_sample", 3.0))
    ladder_window = qc_settings.get("ladder_peak_window_bp", qc_settings.get("w_ladder", 3.0))
    return QCRules(
        min_r2_ok=qc_settings.get("min_r2_ok", 0.999),
        min_r2_warn=qc_settings.get("min_r2_warn", 0.995),
        nk_ymax_floor=qc_settings.get("nk_ymax_floor", 250.0),
        sample_peak_window_bp=sample_window,
        sample_peak_window_bp_fallback=qc_settings.get("sample_peak_window_bp_fallback", max(float(sample_window) + 4.0, 8.0)),
        ladder_peak_window_bp=ladder_window,
        min_sl_total_area=qc_settings.get("min_sl_total_area", 1e4),
    )


def resolve_original_input_path(path_like: Path | str | None) -> Path | None:
    if not path_like:
        return None

    path = Path(path_like).expanduser()
    try:
        if path.is_symlink():
            return path.resolve(strict=False)
    except OSError:
        return None

    try:
        resolved = path.resolve(strict=False)
    except OSError:
        resolved = path

    if path.parent.name.startswith("fraggler_stage_") and resolved.parent == path.parent:
        return None
    return resolved


def resolve_source_run_dir(entry: dict) -> str:
    existing = str(entry.get("source_run_dir") or "").strip()
    if existing:
        return existing

    fsa = entry.get("fsa")
    original_path = resolve_original_input_path(getattr(fsa, "file", None))
    if original_path is not None and original_path.parent.name:
        return original_path.parent.name

    file_name = str(getattr(fsa, "file_name", "") or entry.get("file_name") or "")
    run_key = make_run_key(file_name)
    if run_key and run_key != "UNKNOWN":
        return run_key

    path = Path(getattr(fsa, "file", "")) if getattr(fsa, "file", None) else None
    if path is not None and path.parent.name and not path.parent.name.startswith("fraggler_stage_"):
        return path.parent.name

    return ""


def update_clonality_tracking_workbook(
    excel_path: Path,
    entries: list[dict],
    rules: QCRules | None = None,
    refresh_dashboard: bool = True,
) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    rules = rules or build_clonality_qc_rules()
    df_runs, df_peaks, pk_identity_keys = _build_tracking_frames(entries, rules)
    # If no data and file exists, nothing to do. If no data and file MISSING, we create the skeleton below.
    if df_runs.empty and df_peaks.empty and excel_path.exists():
        return

    with _clonality_excel_lock:
        if excel_path.exists():
            try:
                with pd.ExcelFile(excel_path, engine="openpyxl") as xls:
                    has_runs = "Runs" in xls.sheet_names
                    has_peaks = "PK_Peaks" in xls.sheet_names
            except Exception:
                has_runs = False
                has_peaks = False
                from fraggler.fraggler import print_warning # Import if not already globally available
                print_warning(f"Kunne ikke lese eksisterende {excel_path.name}, kanskje korrupt. Lager ny...")

            try:
                old_runs = pd.read_excel(excel_path, sheet_name="Runs", engine="openpyxl") if has_runs else pd.DataFrame(columns=RUN_SHEET_COLUMNS)
                old_peaks = pd.read_excel(excel_path, sheet_name="PK_Peaks", engine="openpyxl") if has_peaks else pd.DataFrame(columns=PEAK_SHEET_COLUMNS)
            except Exception:
                old_runs = pd.DataFrame(columns=RUN_SHEET_COLUMNS)
                old_peaks = pd.DataFrame(columns=PEAK_SHEET_COLUMNS)
        else:
            old_runs = pd.DataFrame(columns=RUN_SHEET_COLUMNS)
            old_peaks = pd.DataFrame(columns=PEAK_SHEET_COLUMNS)

        old_runs = _normalize_run_frame(old_runs)
        old_peaks = _reindex_columns(old_peaks, PEAK_SHEET_COLUMNS)

        if not df_runs.empty and "IdentityKey" in old_runs.columns:
            old_runs = old_runs[~old_runs["IdentityKey"].isin(df_runs["IdentityKey"])]
        if pk_identity_keys and "IdentityKey" in old_peaks.columns:
            old_peaks = old_peaks[~old_peaks["IdentityKey"].isin(sorted(pk_identity_keys))]

        all_runs = _concat_frames(old_runs, df_runs)
        all_peaks = _concat_frames(old_peaks, df_peaks)

        if not all_runs.empty and "IdentityKey" in all_runs.columns:
            all_runs = all_runs.drop_duplicates(subset=["IdentityKey"], keep="last")
        if not all_peaks.empty and {"IdentityKey", "MarkerName"}.issubset(all_peaks.columns):
            all_peaks = all_peaks.drop_duplicates(subset=["IdentityKey", "MarkerName"], keep="last")

        all_runs = _reindex_columns(all_runs, RUN_SHEET_COLUMNS)
        all_peaks = _reindex_columns(all_peaks, PEAK_SHEET_COLUMNS)

        writer_kwargs = {"engine": "openpyxl"}
        if excel_path.exists():
            writer_kwargs.update({"mode": "a", "if_sheet_exists": "replace"})

        with pd.ExcelWriter(excel_path, **writer_kwargs) as writer:
            all_runs.to_excel(writer, sheet_name="Runs", index=False)
            all_peaks.to_excel(writer, sheet_name="PK_Peaks", index=False)
        _apply_reference_tracking_headers(excel_path)
        if refresh_dashboard:
            refresh_clonality_tracking_dashboard(excel_path)
        print_green(f"Clonality tracking workbook updated in {excel_path}")


def _build_tracking_frames(entries: list[dict], rules: QCRules) -> tuple[pd.DataFrame, pd.DataFrame, set[str]]:
    run_rows: list[dict] = []
    peak_rows: list[dict] = []
    pk_identity_keys: set[str] = set()

    for entry in entries:
        base_row = _build_run_row(entry)
        if not base_row:
            continue

        run_rows.append(base_row)

        if base_row["SampleKind"] == "control" and base_row["Control"] in {"PK", "PK1", "PK2"}:
            pk_identity_keys.add(base_row["IdentityKey"])
            peak_rows.extend(_build_pk_peak_rows(entry, rules, base_row))

    return (
        _reindex_columns(pd.DataFrame(run_rows), RUN_SHEET_COLUMNS),
        _reindex_columns(pd.DataFrame(peak_rows), PEAK_SHEET_COLUMNS),
        pk_identity_keys,
    )


def build_tracking_join_fields(entry: dict) -> dict[str, str]:
    file_name = _resolve_entry_file_name(entry)
    if not file_name:
        return {}

    source_run_dir = resolve_source_run_dir(entry)
    control = control_id_from_filename(file_name)
    is_control = control in CONTROL_IDS
    identity_key = (
        f"{source_run_dir}::{file_name}" if is_control else _build_patient_identity_key(source_run_dir, file_name)
    )
    return {
        "identity_key": identity_key,
        "file_name": file_name,
        "source_run_dir": source_run_dir,
        "assay": str(entry.get("assay") or ""),
        "sample_kind": "control" if is_control else "patient",
        "group": str(entry.get("group") or ""),
        "control": control if is_control else "",
        "run_date": parse_pcr_date_from_filename(file_name) or "",
        "run_code": parse_run_code_from_filename(file_name) or "",
        "well": parse_well_from_filename(file_name) or "",
        "batch": parse_batch_from_filename(file_name) or "",
        "dit": str(entry.get("dit") or ""),
        "ladder": str(entry.get("ladder") or ""),
    }


def build_tracking_join_key(
    *,
    identity_key: str,
    assay: str,
    run_code: str,
    well: str,
) -> str:
    return "::".join(str(value or "").strip() for value in (identity_key, assay, well, run_code))


def build_tracking_ladder_join_key(
    *,
    identity_key: str,
    assay: str,
    run_code: str,
    well: str,
    ladder: str,
) -> str:
    return "::".join(
        str(value or "").strip()
        for value in (build_tracking_join_key(identity_key=identity_key, assay=assay, run_code=run_code, well=well), ladder)
    )


def build_tracking_pk_join_key(
    *,
    identity_key: str,
    assay: str,
    run_code: str,
    well: str,
    marker_name: str,
) -> str:
    return "::".join(
        str(value or "").strip()
        for value in (
            build_tracking_join_key(identity_key=identity_key, assay=assay, run_code=run_code, well=well),
            marker_name,
        )
    )


def build_tracking_row_key(*, artifact_kind: str, identity_key: str, marker_name: str = "") -> str:
    identity = str(identity_key or "").strip()
    if artifact_kind == "pk":
        return f"pk:{identity}:{str(marker_name or '').strip()}"
    if artifact_kind == "ladder":
        return f"ladder:{identity}"
    return identity


def _build_run_row(entry: dict) -> dict:
    join_fields = build_tracking_join_fields(entry)
    if not join_fields:
        return {}
    ladder_r2 = entry.get("ladder_r2")
    if ladder_r2 is None or not np.isfinite(ladder_r2):
        ladder_r2 = ""
    ladder_mean_residual_bp = entry.get("ladder_mean_residual_bp")
    if ladder_mean_residual_bp is None or not np.isfinite(ladder_mean_residual_bp):
        ladder_mean_residual_bp = ""
    ladder_max_residual_bp = entry.get("ladder_max_residual_bp")
    if ladder_max_residual_bp is None or not np.isfinite(ladder_max_residual_bp):
        ladder_max_residual_bp = ""
    ladder_linear_r2 = entry.get("ladder_linear_r2")
    if ladder_linear_r2 is None or not np.isfinite(ladder_linear_r2):
        ladder_linear_r2 = ""
    ladder_quadratic_r2 = entry.get("ladder_quadratic_r2")
    if ladder_quadratic_r2 is None or not np.isfinite(ladder_quadratic_r2):
        ladder_quadratic_r2 = ""
    ladder_linear_mean_residual_bp = entry.get("ladder_linear_mean_residual_bp")
    if ladder_linear_mean_residual_bp is None or not np.isfinite(ladder_linear_mean_residual_bp):
        ladder_linear_mean_residual_bp = ""
    ladder_linear_max_residual_bp = entry.get("ladder_linear_max_residual_bp")
    if ladder_linear_max_residual_bp is None or not np.isfinite(ladder_linear_max_residual_bp):
        ladder_linear_max_residual_bp = ""
    ladder_quadratic_mean_residual_bp = entry.get("ladder_quadratic_mean_residual_bp")
    if ladder_quadratic_mean_residual_bp is None or not np.isfinite(ladder_quadratic_mean_residual_bp):
        ladder_quadratic_mean_residual_bp = ""
    ladder_quadratic_max_residual_bp = entry.get("ladder_quadratic_max_residual_bp")
    if ladder_quadratic_max_residual_bp is None or not np.isfinite(ladder_quadratic_max_residual_bp):
        ladder_quadratic_max_residual_bp = ""
    ladder_curvature = entry.get("ladder_max_curvature")
    if ladder_curvature is None or not np.isfinite(ladder_curvature):
        ladder_curvature = ""
    rust_preview_top_score = entry.get("rust_preview_top_score")
    if rust_preview_top_score is None or not np.isfinite(rust_preview_top_score):
        rust_preview_top_score = ""
    rust_preview_top_dominant_ratio = entry.get("rust_preview_top_dominant_ratio")
    if rust_preview_top_dominant_ratio is None or not np.isfinite(rust_preview_top_dominant_ratio):
        rust_preview_top_dominant_ratio = ""

    return {
        "Month": _month_bucket(join_fields["run_date"]),
        "IdentityKey": join_fields["identity_key"],
        "File": join_fields["file_name"],
        "SourceRunDir": join_fields["source_run_dir"],
        "DIT": join_fields["dit"],
        "Assay": join_fields["assay"],
        "SampleKind": join_fields["sample_kind"],
        "Group": join_fields["group"],
        "Control": join_fields["control"],
        "RunDate": join_fields["run_date"],
        "RunCode": join_fields["run_code"],
        "Well": join_fields["well"],
        "Batch": join_fields["batch"],
        "Ladder": join_fields["ladder"],
        "LadderQC": entry.get("ladder_qc_status") or "",
        "LadderFitStrategy": entry.get("ladder_fit_strategy") or "",
        "LadderExpectedStepCount": int(entry.get("ladder_expected_step_count", 0) or 0),
        "LadderFittedStepCount": int(entry.get("ladder_fitted_step_count", 0) or 0),
        "LadderR2": ladder_r2,
        "LadderLinearR2": ladder_linear_r2,
        "LadderLinearMeanResidualBp": ladder_linear_mean_residual_bp,
        "LadderLinearMaxResidualBp": ladder_linear_max_residual_bp,
        "LadderCurvature": ladder_curvature,
    }


def _build_pk_peak_rows(entry: dict, rules: QCRules, base_row: dict) -> list[dict]:
    marker_rows: list[dict] = []
    fsa = entry["fsa"]
    primary_channel = str(entry.get("primary_peak_channel") or "")

    tracking_results = entry.get("tracking_marker_results")
    for marker in markers_for_entry(entry, rules):
        channel = primary_channel if marker["channel"] == "primary" else str(marker["channel"])

        result = None
        if tracking_results and marker["name"] in tracking_results:
            result = tracking_results[marker["name"]]

        if result is None:
            if entry.get("fsa") is None:
                result = {"ok": False, "reason": "FSA data dropped and no pre-calculated results"}
            elif marker["kind"] == "sample":
                result = find_peak_near_bp_with_fallback(
                    fsa=fsa,
                    channel=channel,
                    target_bp=float(marker["expected_bp"]),
                    window_bp=float(marker["window_bp"]),
                    fallback_window_bp=float(getattr(rules, "sample_peak_window_bp_fallback", marker["window_bp"])),
                    baseline_correct=True,
                )
            else:
                result = find_peak_near_bp(
                    fsa=fsa,
                    channel=channel,
                    target_bp=float(marker["expected_bp"]),
                    window_bp=float(marker["window_bp"]),
                    baseline_correct=True,
                )

        row = {
            "Month": _month_bucket(base_row["RunDate"]),
            "IdentityKey": base_row["IdentityKey"],
            "File": base_row["File"],
            "SourceRunDir": base_row["SourceRunDir"],
            "DIT": base_row["DIT"],
            "Assay": base_row["Assay"],
            "Control": base_row["Control"],
            "RunDate": base_row["RunDate"],
            "RunCode": base_row["RunCode"],
            "Well": base_row["Well"],
            "Batch": base_row["Batch"],
            "MarkerName": marker.get("name") or "",
            "Kind": marker.get("kind") or "",
            "Channel": channel,
            "ExpectedBP": float(marker.get("expected_bp", np.nan)),
            "WindowBP": float(marker.get("window_bp", np.nan)),
            "SearchMode": result.get("search_mode", ""),
            "SearchWindowBP": float(result.get("search_window_bp", np.nan)) if result.get("search_window_bp") is not None else "",
            "FoundBP": "",
            "DeltaBP": "",
            "Height": "",
            "Area": "",
            "OK": bool(result.get("ok", False)),
            "Reason": result.get("reason") or "",
            "AbsDeltaBP": "",
        }
        if result.get("ok", False):
            found_bp = float(result["found_bp"])
            delta_bp = found_bp - float(marker.get("expected_bp", np.nan))
            row["FoundBP"] = found_bp
            row["DeltaBP"] = delta_bp
            row["Height"] = float(result.get("height", np.nan))
            row["Area"] = float(result.get("area", np.nan))
            row["AbsDeltaBP"] = abs(delta_bp)
        marker_rows.append(row)

    return marker_rows


def _reindex_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=columns)
    return df.reindex(columns=columns)


def _concat_frames(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    frames = [df for df in (old_df, new_df) if not df.empty]
    if not frames:
        return pd.DataFrame(columns=(list(old_df.columns) if len(old_df.columns) else list(new_df.columns)))
    if len(frames) == 1:
        return frames[0].copy()
    return pd.concat(frames, ignore_index=True)


def _resolve_entry_file_name(entry: dict) -> str:
    fsa = entry.get("fsa")
    original_path = resolve_original_input_path(getattr(fsa, "file", None))
    if original_path is not None and original_path.name:
        return original_path.name
    return strip_stage_prefix(str(getattr(fsa, "file_name", "") or entry.get("file_name") or ""))


def sanitize_clonality_tracking_workbook(excel_path: Path, *, refresh_dashboard: bool = True) -> bool:
    excel_path = Path(excel_path).expanduser()
    if not excel_path.exists():
        return False

    with _clonality_excel_lock:
        with pd.ExcelFile(excel_path, engine="openpyxl") as xls:
            has_runs = "Runs" in xls.sheet_names
            has_peaks = "PK_Peaks" in xls.sheet_names
        if not any([has_runs, has_peaks]):
            return False

        runs = pd.read_excel(excel_path, sheet_name="Runs", engine="openpyxl") if has_runs else pd.DataFrame(columns=RUN_SHEET_COLUMNS)
        peaks = pd.read_excel(excel_path, sheet_name="PK_Peaks", engine="openpyxl") if has_peaks else pd.DataFrame(columns=PEAK_SHEET_COLUMNS)

        runs = _normalize_run_frame(runs)
        peaks = _reindex_columns(peaks, PEAK_SHEET_COLUMNS)

        writer_kwargs = {"engine": "openpyxl", "mode": "a", "if_sheet_exists": "replace"}
        with pd.ExcelWriter(excel_path, **writer_kwargs) as writer:
            runs.to_excel(writer, sheet_name="Runs", index=False)
            peaks.to_excel(writer, sheet_name="PK_Peaks", index=False)
        _apply_reference_tracking_headers(excel_path)

        if refresh_dashboard:
            refresh_clonality_tracking_dashboard(excel_path)
    return True


def _normalize_run_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=RUN_SHEET_COLUMNS)

    normalized = df.copy()
    legacy_identity = normalized.get("IdentityKey", pd.Series("", index=normalized.index)).fillna("").astype(str)
    source_run_dir = normalized.get("SourceRunDir", pd.Series("", index=normalized.index)).fillna("").astype(str)
    file_name = normalized.get("File", pd.Series("", index=normalized.index)).fillna("").astype(str)
    source_run_dir = source_run_dir.where(source_run_dir.str.strip() != "", legacy_identity.map(_legacy_source_run_dir_from_identity_key))
    file_name = file_name.where(file_name.str.strip() != "", legacy_identity.map(_legacy_file_name_from_identity_key))
    normalized_identity: list[str] = []
    
    # We should distinguish if this is a patient run by looking at Control column or SampleKind
    sample_kind = normalized.get("SampleKind", pd.Series("", index=normalized.index)).fillna("").astype(str)
    control = normalized.get("Control", pd.Series("", index=normalized.index)).fillna("").astype(str)

    for src, fname, legacy, skind, ctrl in zip(source_run_dir.tolist(), file_name.tolist(), legacy_identity.tolist(), sample_kind.tolist(), control.tolist()):
        legacy_value = str(legacy or "").strip()
        is_control = ctrl in CONTROL_IDS or skind == "control"
        
        if is_control:
            normalized_identity.append(f"{src}::{fname}")
        else:
            if legacy_value.startswith("PT-") and not str(fname or "").strip().lower().endswith(".fsa"):
                normalized_identity.append(legacy_value)
                continue
            normalized_identity.append(_build_patient_identity_key(src, fname or legacy))
            
    normalized["IdentityKey"] = normalized_identity
    normalized["SourceRunDir"] = source_run_dir
    return _reindex_columns(normalized, RUN_SHEET_COLUMNS)


def _build_patient_identity_key(source_run_dir: str, file_name: str) -> str:
    identity_source = f"{str(source_run_dir or '').strip()}::{str(file_name or '').strip()}"
    if not identity_source.strip(":"):
        identity_source = str(file_name or source_run_dir or "UNKNOWN_PATIENT")
    digest = hmac.new(
        _get_tracking_identity_salt().encode("utf-8"),
        identity_source.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"PT-{digest[:24]}"


def _get_tracking_identity_salt() -> str:
    batch_settings = APP_SETTINGS.get("analyses", {}).get("clonality", {}).get("batch", {})
    configured = str(batch_settings.get("tracking_identity_salt") or "").strip()
    if configured:
        return configured

    env_salt = str(os.environ.get(TRACKING_IDENTITY_SALT_ENV, "") or "").strip()
    if env_salt:
        return env_salt

    salt_path = str(os.environ.get(TRACKING_IDENTITY_SALT_PATH_ENV, "") or batch_settings.get("tracking_identity_salt_path") or "").strip()
    path = Path(salt_path).expanduser() if salt_path else DEFAULT_TRACKING_IDENTITY_SALT_PATH
    try:
        if path.exists():
            existing = path.read_text(encoding="utf-8", errors="replace").strip()
            if existing:
                return existing
        path.parent.mkdir(parents=True, exist_ok=True)
        generated = secrets.token_hex(32)
        path.write_text(generated, encoding="utf-8")
        return generated
    except OSError:
        # Last-resort fallback keeps behavior deterministic within a machine/user context.
        return hashlib.sha256(str(path).encode("utf-8")).hexdigest()


def _legacy_source_run_dir_from_identity_key(identity_key: str) -> str:
    value = str(identity_key or "")
    if "::" not in value:
        return ""
    return value.split("::", 1)[0].strip()


def _legacy_file_name_from_identity_key(identity_key: str) -> str:
    value = str(identity_key or "")
    if "::" in value:
        return value.split("::", 1)[1].strip()
    return value.strip()


def _month_bucket(run_date: str) -> str:
    value = str(run_date or "").strip()
    if len(value) >= 7 and value[4:5] == "-" and value[7:8] == "-":
        return value[:7].replace("-", "_")
    return ""


def _apply_reference_tracking_headers(excel_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    if "Runs" in wb.sheetnames:
        ws = wb["Runs"]
        for col, header in enumerate(RUN_SHEET_COLUMNS, start=1):
            ws.cell(1, col).value = header
    wb.save(excel_path)
