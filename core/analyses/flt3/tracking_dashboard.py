from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="174A5B")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF0")
CARD_GREEN = PatternFill("solid", fgColor="E2F0D9")
CARD_BLUE = PatternFill("solid", fgColor="DDEBF7")
CARD_GOLD = PatternFill("solid", fgColor="FFF2CC")
CARD_RED = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D8E1E5")
BOX_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)
ACCEPTED_LADDER_STATUSES = {"ok", "manual_adjustment"}
GENERATED_SHEETS = {
    "Dashboard",
    "Dashboard_Data",
    "Assay_Summary",
    "Run_Summary",
    "Control_Summary",
    "PK_Sample_Delta",
    "PK_Ladder_Delta",
    "FLT3_Assay_Summary",
    "FLT3_Run_Summary",
    "FLT3_Control_Summary",
    "FLT3_Ratio_Trends",
    "FLT3_Marker_QC",
    "QC_Run_Trends",
    "QC_Control_Signals",
    "QC_Baseline_Config",
    "Review_Rows",
    "Summary_By_Injection",
}
OBSOLETE_TRACKING_COLUMNS = {
    "LadderEngine",
    "LadderReasonCodes",
    "SourceFsaSha256",
    "ManualAdjustmentSha256",
    "AnalysisVersion",
    "PullUpCandidate",
    "SaturationCandidate",
}


def refresh_flt3_tracking_dashboard(excel_path: Path) -> None:
    if not excel_path.exists():
        return

    with pd.ExcelFile(excel_path, engine="openpyxl") as xls:
        sheet_names = set(xls.sheet_names)
        required = {"Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"}
        frames = (
            {
                name: pd.read_excel(
                    excel_path,
                    sheet_name=name,
                    engine="openpyxl",
                )
                for name in required
            }
            if required.issubset(sheet_names)
            else {}
        )

    wb = load_workbook(excel_path)
    try:
        for name in list(wb.sheetnames):
            if name in GENERATED_SHEETS:
                del wb[name]

        for sheet_name in (
            "Runs",
            "Patient_Runs",
            "Control_Runs",
            "PK_Peaks",
            "All_Analyzed_QC",
            "Raw_Metadata_All_FSA",
            "Skipped",
        ):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            _remove_obsolete_columns(ws)
            if sheet_name in {"Runs", "Patient_Runs", "Control_Runs"}:
                _ensure_manual_adjustment_column(ws)
            _style_data_sheet(ws)
        for sheet_name in ("Runs", "Patient_Runs", "Control_Runs"):
            if sheet_name in wb.sheetnames:
                _add_review_validation(wb[sheet_name])

        if frames:
            runs = frames["Runs"]
            assay_summary = _build_assay_summary(runs)
            per_run_summary = _build_per_run_summary(runs)
            control_summary = _build_control_summary(runs)
            dashboard = wb.create_sheet("Dashboard", 0)
            _build_dashboard(
                dashboard,
                runs=runs,
                assay_summary=assay_summary,
                per_run_summary=per_run_summary,
                control_summary=control_summary,
            )
            _add_dashboard_charts(
                dashboard,
                assay_count=min(len(assay_summary), 12),
            )

        wb.save(excel_path)
    finally:
        wb.close()


def _text(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series("", index=frame.index, dtype="object")
    return frame[column].fillna("").astype(str).str.strip()


def _numeric(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(np.nan, index=frame.index, dtype=float)
    return pd.to_numeric(frame[column], errors="coerce")


def _boolean(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(False, index=frame.index, dtype=bool)
    return _text(frame, column).str.lower().isin({"true", "1", "yes", "y"})


def _review_mask(series: pd.Series, pass_statuses: set[str]) -> pd.Series:
    values = series.fillna("").astype(str).str.strip().str.lower()
    return values.ne("") & ~values.isin(pass_statuses)


def _run_keys(frame: pd.DataFrame) -> pd.Series:
    source = _text(frame, "SourceRunDir")
    code = _text(frame, "RunCode")
    date = _text(frame, "RunDate")
    file_name = _text(frame, "File")
    keys = source.where(source.ne(""), code)
    keys = keys.where(keys.ne(""), date)
    return keys.where(keys.ne(""), file_name)


def _prepared_runs(runs: pd.DataFrame) -> pd.DataFrame:
    work = runs.copy()
    work["_Assay"] = _text(work, "Assay").replace("", "Unknown")
    work["_RunKey"] = _run_keys(work)
    work["_RunDate"] = _text(work, "RunDate")
    work["_Patient"] = _text(work, "SampleKind").str.lower().eq("patient")
    work["_Control"] = (
        _text(work, "SampleKind").str.lower().eq("control")
        | _text(work, "Control").ne("")
    )
    work["_LadderStatus"] = _text(work, "LadderQC").str.lower()
    work["_Manual"] = (
        work["_LadderStatus"].eq("manual_adjustment")
        | _text(work, "LadderFitStrategy").str.lower().eq("manual_adjustment")
    )
    work["_LadderReview"] = _review_mask(
        work["_LadderStatus"],
        ACCEPTED_LADDER_STATUSES,
    )
    work["_PeakReview"] = _review_mask(
        _text(work, "PeakQC"),
        {"ok", "negative_control", "not_evaluated_ladder_only"},
    )
    work["_Positive"] = _boolean(work, "PositiveCall") | _text(
        work,
        "Interpretation",
    ).str.lower().str.startswith("positiv ")
    work["_Ratio"] = _numeric(work, "Ratio")
    work["_RatioAvailable"] = work["_Ratio"].notna() & ~_text(
        work,
        "RatioMode",
    ).str.lower().eq("manual_required")
    work["_LadderR2"] = _numeric(work, "LadderR2")
    return work


def _build_assay_summary(runs: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Assay",
        "Injections",
        "PatientInjections",
        "ControlInjections",
        "ReviewRequired",
        "ManualAdjusted",
        "PositiveCalls",
        "MeanLadderR2",
    ]
    work = _prepared_runs(runs)
    rows = []
    for assay, group in work.groupby("_Assay", sort=True):
        rows.append(
            {
                "Assay": assay,
                "Injections": len(group),
                "PatientInjections": int(group["_Patient"].sum()),
                "ControlInjections": int(group["_Control"].sum()),
                "ReviewRequired": int(
                    (group["_LadderReview"] | group["_PeakReview"]).sum()
                ),
                "ManualAdjusted": int(group["_Manual"].sum()),
                "PositiveCalls": int(group["_Positive"].sum()),
                "MeanLadderR2": group["_LadderR2"].mean(),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _build_per_run_summary(runs: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "RunDate",
        "Run",
        "Injections",
        "Patients",
        "Controls",
        "ReviewRequired",
        "ManualAdjusted",
    ]
    work = _prepared_runs(runs)
    work = work.loc[work["_RunKey"].ne("")]
    rows = []
    for (run_date, run_key), group in work.groupby(
        ["_RunDate", "_RunKey"],
        sort=True,
        dropna=False,
    ):
        rows.append(
            {
                "RunDate": run_date,
                "Run": run_key,
                "Injections": len(group),
                "Patients": int(group["_Patient"].sum()),
                "Controls": int(group["_Control"].sum()),
                "ReviewRequired": int(
                    (group["_LadderReview"] | group["_PeakReview"]).sum()
                ),
                "ManualAdjusted": int(group["_Manual"].sum()),
            }
        )
    return pd.DataFrame(rows, columns=columns).sort_values(
        ["RunDate", "Run"],
        kind="stable",
        ignore_index=True,
    )


def _build_control_summary(runs: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Control",
        "Assay",
        "Injections",
        "ReviewRequired",
        "ManualAdjusted",
        "PositiveCalls",
        "MeanWTBP",
        "MeanMutantBP",
        "MeanRatio",
    ]
    work = _prepared_runs(runs)
    work["_ControlCode"] = _text(work, "Control")
    work["_WTBP"] = _numeric(work, "WT_BP")
    work["_MutantBP"] = _numeric(work, "MutantMain_BP")
    work = work.loc[work["_ControlCode"].ne("")]
    rows = []
    for (control, assay), group in work.groupby(
        ["_ControlCode", "_Assay"],
        sort=True,
    ):
        rows.append(
            {
                "Control": control,
                "Assay": assay,
                "Injections": len(group),
                "ReviewRequired": int(
                    (group["_LadderReview"] | group["_PeakReview"]).sum()
                ),
                "ManualAdjusted": int(group["_Manual"].sum()),
                "PositiveCalls": int(group["_Positive"].sum()),
                "MeanWTBP": group["_WTBP"].mean(),
                "MeanMutantBP": group["_MutantBP"].mean(),
                "MeanRatio": group["_Ratio"].mean(),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _build_dashboard(
    ws,
    *,
    runs: pd.DataFrame,
    assay_summary: pd.DataFrame,
    per_run_summary: pd.DataFrame,
    control_summary: pd.DataFrame,
) -> None:
    work = _prepared_runs(runs)
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, 18):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["K"].width = 25

    ws.merge_cells("A1:Q2")
    ws["A1"] = "HemaFrag FLT3 Tracking"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A3:Q3")
    ws["A3"] = (
        "Historical run ledger: new injections append; re-analysis updates "
        "the existing run."
    )
    ws["A3"].font = Font(color="52666D", italic=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    latest_date = _text(work, "RunDate").max() if not work.empty else ""
    metrics = [
        ("Tracked injections", len(work), CARD_BLUE),
        ("Unique runs", int(_run_keys(work).replace("", np.nan).nunique()), CARD_BLUE),
        ("Patient injections", int(work["_Patient"].sum()), CARD_GREEN),
        ("Control injections", int(work["_Control"].sum()), CARD_GREEN),
        (
            "Needs review",
            int((work["_LadderReview"] | work["_PeakReview"]).sum()),
            CARD_RED,
        ),
        ("Manual adjusted", int(work["_Manual"].sum()), CARD_GOLD),
        ("Positive calls", int(work["_Positive"].sum()), CARD_GOLD),
        ("Latest run", latest_date, CARD_BLUE),
    ]
    for index, (label, value, fill) in enumerate(metrics, start=1):
        ws.cell(5, index, label)
        ws.cell(6, index, value)
        for row in (5, 6):
            cell = ws.cell(row, index)
            cell.fill = fill
            cell.border = BOX_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(5, index).font = BOLD

    ws["J8"] = "Recent runs"
    ws["J8"].font = Font(size=13, bold=True)
    recent = per_run_summary.tail(10).sort_values(
        ["RunDate", "Run"],
        ascending=False,
        kind="stable",
    )
    _write_dashboard_frame(ws, recent, start_row=9, start_col=10)

    ws["A15"] = "Assay Watchlist"
    ws["A15"].font = Font(size=13, bold=True)
    _write_dashboard_frame(
        ws,
        assay_summary.head(12),
        start_row=16,
        start_col=1,
    )

    control_start = max(31, 18 + len(assay_summary.head(12)))
    ws.cell(control_start, 1, "Control overview")
    ws.cell(control_start, 1).font = Font(size=13, bold=True)
    _write_dashboard_frame(
        ws,
        control_summary.head(12),
        start_row=control_start + 1,
        start_col=1,
    )
    ws.freeze_panes = "A9"


def _write_dashboard_frame(
    ws,
    frame: pd.DataFrame,
    *,
    start_row: int,
    start_col: int,
) -> None:
    for col_offset, header in enumerate(frame.columns):
        ws.cell(start_row, start_col + col_offset, header)
    for row_offset, row in enumerate(
        frame.itertuples(index=False, name=None),
        start=1,
    ):
        for col_offset, value in enumerate(row):
            ws.cell(
                start_row + row_offset,
                start_col + col_offset,
                None if pd.isna(value) else value,
            )
    _style_range(
        ws,
        start_row,
        start_row + max(len(frame), 1),
        start_col,
        start_col + max(len(frame.columns), 1) - 1,
    )


def _add_dashboard_charts(ws, *, assay_count: int) -> None:
    if not assay_count:
        return
    volume_chart = BarChart()
    volume_chart.type = "col"
    volume_chart.style = 10
    volume_chart.title = "Patient and control injections by assay"
    volume_chart.y_axis.title = "Injections"
    volume_chart.height = 7
    volume_chart.width = 12
    volume_chart.add_data(
        Reference(
            ws,
            min_col=3,
            max_col=4,
            min_row=16,
            max_row=16 + assay_count,
        ),
        titles_from_data=True,
    )
    volume_chart.set_categories(
        Reference(ws, min_col=1, min_row=17, max_row=16 + assay_count)
    )
    volume_chart.series[0].tx = SeriesLabel(v="Patient")
    volume_chart.series[1].tx = SeriesLabel(v="Control")
    volume_chart.series[0].graphicalProperties.solidFill = "70AD8F"
    volume_chart.series[1].graphicalProperties.solidFill = "4F81BD"
    ws.add_chart(volume_chart, "A47")

    review_chart = BarChart()
    review_chart.type = "col"
    review_chart.style = 11
    review_chart.title = "Review and manual adjustments by assay"
    review_chart.y_axis.title = "Injections"
    review_chart.height = 7
    review_chart.width = 12
    review_chart.add_data(
        Reference(
            ws,
            min_col=5,
            max_col=6,
            min_row=16,
            max_row=16 + assay_count,
        ),
        titles_from_data=True,
    )
    review_chart.set_categories(
        Reference(ws, min_col=1, min_row=17, max_row=16 + assay_count)
    )
    review_chart.series[0].tx = SeriesLabel(v="Review required")
    review_chart.series[1].tx = SeriesLabel(v="Manual adjusted")
    review_chart.series[0].graphicalProperties.solidFill = "C0504D"
    review_chart.series[1].graphicalProperties.solidFill = "E5B84B"
    ws.add_chart(review_chart, "J47")


def _remove_obsolete_columns(ws) -> None:
    for column_index in range(ws.max_column, 0, -1):
        header = str(ws.cell(1, column_index).value or "").strip()
        if (
            header in OBSOLETE_TRACKING_COLUMNS
            or header.startswith("GS500ROXStartPrior")
        ):
            ws.delete_cols(column_index)


def _ensure_manual_adjustment_column(ws) -> None:
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    if "ManualAdjustmentUsed" in headers:
        return
    strategy_col = headers.get("LadderFitStrategy")
    if not strategy_col:
        return
    target_col = strategy_col + 1
    ws.insert_cols(target_col)
    ws.cell(1, target_col, "ManualAdjustmentUsed")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(
            row_idx,
            target_col,
            str(ws.cell(row_idx, strategy_col).value or "").strip().lower()
            == "manual_adjustment",
        )


def _style_range(
    ws,
    header_row: int,
    data_end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(header_row, col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX_BORDER
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=max(data_end_row, header_row + 1),
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = BOX_BORDER
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFB")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def _style_data_sheet(ws) -> None:
    if ws.max_column < 1:
        return
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = BOX_BORDER
    ws.row_dimensions[1].height = 30
    if ws.title in {"Runs", "Patient_Runs", "Control_Runs"}:
        ws.freeze_panes = "G2"
    elif ws.title == "PK_Peaks":
        ws.freeze_panes = "F2"
    else:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    identity_col = headers.get("IdentityKey")
    if identity_col:
        ws.column_dimensions[get_column_letter(identity_col)].hidden = True

    sample_end = min(ws.max_row, 200)
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        width = len(header) + 2
        for row_idx in range(2, sample_end + 1):
            width = max(width, len(str(ws.cell(row_idx, col_idx).value or "")) + 1)
        cap = 42 if header in {
            "File",
            "SourceRunDir",
            "Interpretation",
            "Reason",
            "TrackingNote",
        } else 24
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width, 10),
            cap,
        )
        normalized = header.lower()
        if "ratio" in normalized or normalized.endswith("fraction"):
            number_format = "0.0000"
        elif "r2" in normalized:
            number_format = "0.000000"
        elif "bp" in normalized or "area" in normalized or "height" in normalized:
            number_format = "0.00"
        else:
            number_format = None
        if number_format:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = number_format

    for status_header in ("LadderQC", "PeakQC", "QCStatus"):
        col_idx = headers.get(status_header)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            value = str(cell.value or "").strip().lower()
            if not value:
                continue
            if value in {
                "ok",
                "pass",
                "manual_adjustment",
                "negative_control",
                "not_evaluated_ladder_only",
            }:
                cell.fill = CARD_GREEN
            elif "review" in value or "manual" in value:
                cell.fill = CARD_GOLD
            else:
                cell.fill = CARD_RED

    ratio_col = headers.get("Ratio")
    if ratio_col and ws.max_row >= 2:
        ratio_range = (
            f"{get_column_letter(ratio_col)}2:"
            f"{get_column_letter(ratio_col)}{ws.max_row}"
        )
        ws.conditional_formatting.add(
            ratio_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=SUBHEADER_FILL,
            ),
        )


def _add_review_validation(ws) -> None:
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    column = headers.get("ReviewStatus")
    if not column:
        return
    validation = DataValidation(
        type="list",
        formula1='"Not reviewed,Reviewed,Follow-up"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    validation_end = max(ws.max_row + 5000, 5000)
    validation.add(
        f"{get_column_letter(column)}2:{get_column_letter(column)}{validation_end}"
    )
