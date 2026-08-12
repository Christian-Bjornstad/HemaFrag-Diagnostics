"""Durable, atomic provenance manifests for HemaFrag batch runs."""
from __future__ import annotations

import hashlib
import json
import os
import platform
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence


RUN_MANIFEST_SCHEMA = "hemafrag_batch_run_manifest_v1"
RUN_MANIFEST_PREFIX = "hemafrag_run_"


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _fingerprint(payload: object) -> str:
    encoded = json.dumps(
        payload,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _atomic_write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
            json.dump(payload, handle, indent=2, ensure_ascii=True, default=str)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def _file_record(path_value: object) -> dict[str, object]:
    path = Path(str(path_value)).expanduser()
    try:
        resolved = path.resolve()
    except Exception:
        resolved = path
    record: dict[str, object] = {
        "path": str(resolved),
        "file_name": resolved.name,
        "exists": resolved.is_file(),
    }
    if resolved.is_file():
        record["size_bytes"] = int(resolved.stat().st_size)
        record["sha256"] = _sha256_file(resolved)
    from core.ladder_adjustment_store import load_ladder_adjustment_record

    adjustment = load_ladder_adjustment_record(resolved)
    if adjustment is not None:
        record["manual_adjustment"] = {
            "storage": "internal",
            "sha256": str(adjustment.get("payload_sha256") or ""),
            "saved_at_utc": str(adjustment.get("saved_at_utc") or ""),
        }
    return record


def _runtime_metadata() -> dict[str, object]:
    try:
        from app_meta import APP_VERSION
    except Exception:
        APP_VERSION = "unknown"
    try:
        from core.rust_bridge import (
            _in_process_native_wheel_is_available,
            _resolve_cli_bin,
        )

        cli_path = _resolve_cli_bin()
        rust_engine: dict[str, object] = {
            "native_wheel_available": bool(_in_process_native_wheel_is_available()),
            "cli_available": cli_path is not None,
        }
        if cli_path is not None and cli_path.is_file():
            rust_engine["cli_sha256"] = _sha256_file(cli_path)
    except Exception as exc:
        rust_engine = {
            "native_wheel_available": False,
            "cli_available": False,
            "probe_error": f"{type(exc).__name__}: {exc}",
        }
    return {
        "app_version": str(APP_VERSION),
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "logical_cpu_count": os.cpu_count(),
        "rust_engine": rust_engine,
    }


def _output_records(output_dir: Path | None) -> list[dict[str, object]]:
    if output_dir is None or not output_dir.is_dir():
        return []
    records: list[dict[str, object]] = []
    for path in sorted(output_dir.rglob("*")):
        if not path.is_file() or path.name.startswith(RUN_MANIFEST_PREFIX):
            continue
        try:
            record: dict[str, object] = {
                "path": str(path.resolve()),
                "relative_path": path.relative_to(output_dir).as_posix(),
                "size_bytes": int(path.stat().st_size),
                "sha256": _sha256_file(path),
            }
            if path.suffix.lower() == ".xlsx":
                try:
                    from openpyxl import load_workbook

                    workbook = load_workbook(path, read_only=True, data_only=False)
                    try:
                        record["sheet_rows"] = {
                            sheet.title: max(0, int(sheet.max_row or 0) - 1)
                            for sheet in workbook.worksheets
                        }
                    finally:
                        workbook.close()
                except Exception as exc:
                    record["sheet_row_probe_error"] = (
                        f"{type(exc).__name__}: {exc}"
                    )
            records.append(record)
        except OSError:
            continue
    return records


class BatchRunManifest:
    """Thread-safe writer for an inspectable batch-run ledger."""

    def __init__(self, path: Path, payload: dict[str, Any]) -> None:
        self.path = path
        self.payload = payload
        self._lock = threading.RLock()
        self._last_write_monotonic = time.monotonic()
        self._job_indexes_by_name: dict[str, list[int]] = {}
        for job in payload.get("jobs", []):
            self._job_indexes_by_name.setdefault(str(job.get("name") or ""), []).append(
                int(job["index"])
            )

    @classmethod
    def create(
        cls,
        *,
        output_dir: Path,
        jobs: Sequence[Mapping[str, Any]],
        analysis: str,
        settings: Mapping[str, Any],
        execution: Mapping[str, Any],
        parent_manifest_path: Path | None = None,
    ) -> "BatchRunManifest":
        output_dir = output_dir.resolve()
        created_at = _utc_now()
        run_id = (
            datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
            + "_"
            + uuid.uuid4().hex[:8]
        )
        job_records: list[dict[str, Any]] = []
        input_count = 0
        for index, job in enumerate(jobs):
            files = [_file_record(path) for path in list(job.get("files") or [])]
            input_count += len(files)
            job_records.append(
                {
                    "index": index,
                    "name": str(job.get("name") or f"job_{index + 1}"),
                    "type": str(job.get("type") or "pipeline"),
                    "source_dir": str(job.get("path") or ""),
                    "status": "pending",
                    "last_phase": "",
                    "files": files,
                    "stages": {},
                }
            )
        payload: dict[str, Any] = {
            "schema_version": RUN_MANIFEST_SCHEMA,
            "run_id": run_id,
            "created_at_utc": created_at,
            "updated_at_utc": created_at,
            "status": "running",
            "analysis": analysis,
            "parent_manifest_path": (
                str(parent_manifest_path.resolve()) if parent_manifest_path else None
            ),
            "runtime": _runtime_metadata(),
            "settings_fingerprint": _fingerprint(settings),
            "execution": dict(execution),
            "counts": {
                "expected_jobs": len(job_records),
                "expected_input_files": input_count,
                "expected_patient_jobs": sum(
                    1 for job in job_records if job["type"] != "qc"
                ),
                "expected_qc_jobs": sum(
                    1 for job in job_records if job["type"] == "qc"
                ),
                "completed_jobs": 0,
                "failed_jobs": 0,
                "dit_entries": 0,
                "qc_entries": 0,
            },
            "jobs": job_records,
            "outputs": {},
        }
        path = output_dir / f"{RUN_MANIFEST_PREFIX}{run_id}.json"
        recorder = cls(path, payload)
        recorder._write()
        return recorder

    def _write(self) -> None:
        self.payload["updated_at_utc"] = _utc_now()
        _atomic_write_json(self.path, self.payload)
        self._last_write_monotonic = time.monotonic()

    def record_progress(self, event: Mapping[str, Any]) -> None:
        name = str(event.get("job_name") or "")
        phase = str(event.get("phase") or "")
        if not name or not phase:
            return
        with self._lock:
            indexes = self._job_indexes_by_name.get(name, [])
            if not indexes:
                return
            job = self.payload["jobs"][indexes[0]]
            previous_phase = str(job.get("last_phase") or "")
            now = _utc_now()
            job["last_phase"] = phase
            job["updated_at_utc"] = now
            job["last_file_name"] = str(event.get("file_name") or "")
            job["files_done"] = event.get("files_done")
            job["files_total"] = event.get("files_total")
            stage = job["stages"].setdefault(
                phase,
                {"first_seen_at_utc": now, "events": 0},
            )
            stage["events"] = int(stage.get("events") or 0) + 1
            stage["last_seen_at_utc"] = now
            if phase == "job_start":
                job["status"] = "running"
            elif phase == "done":
                job["status"] = "completed"
            elif phase == "failed":
                job["status"] = "failed"
                job["error"] = str(event.get("note") or "")
            if (
                phase != previous_phase
                or phase in {"done", "failed"}
                or time.monotonic() - self._last_write_monotonic >= 5.0
            ):
                self._write()

    def finalize(
        self,
        *,
        result: Mapping[str, Any],
        aggregate_output_dir: Path | None,
        review_gate: Mapping[str, Any] | None,
    ) -> None:
        with self._lock:
            completed = {str(value) for value in result.get("completed_jobs") or []}
            failed = {str(value) for value in result.get("failed_jobs") or []}
            result_entries = list(result.get("dit_report_entries") or [])
            entries_by_path: dict[Path, Mapping[str, Any]] = {}
            for entry in result_entries:
                if not isinstance(entry, Mapping):
                    continue
                fsa = entry.get("fsa")
                raw_path = (
                    entry.get("original_file_path")
                    or getattr(fsa, "file", None)
                )
                if raw_path:
                    entries_by_path[Path(str(raw_path)).expanduser().resolve()] = entry
            for job in self.payload.get("jobs", []):
                name = str(job.get("name") or "")
                if name in failed:
                    job["status"] = "failed"
                elif name in completed:
                    job["status"] = "completed"

                for file_record in job.get("files", []):
                    source_path = Path(str(file_record.get("path") or ""))
                    from core.ladder_adjustment_store import (
                        load_ladder_adjustment_record,
                    )

                    adjustment = load_ladder_adjustment_record(source_path)
                    if adjustment is not None:
                        adjustment_record = {
                            "storage": "internal",
                            "sha256": str(
                                adjustment.get("payload_sha256") or ""
                            ),
                            "saved_at_utc": str(
                                adjustment.get("saved_at_utc") or ""
                            ),
                        }
                        entry = entries_by_path.get(source_path.expanduser().resolve())
                        provenance = (
                            entry.get("analysis_provenance")
                            if isinstance(entry, Mapping)
                            and isinstance(entry.get("analysis_provenance"), Mapping)
                            else {}
                        )
                        consumed_hash = str(
                            provenance.get("manual_adjustment_sha256") or ""
                        )
                        consumed = bool(
                            provenance.get("manual_adjustment_consumed")
                            and consumed_hash == adjustment_record["sha256"]
                        )
                        adjustment_record["consumed"] = consumed
                        adjustment_record["consumed_sha256"] = (
                            consumed_hash if consumed else ""
                        )
                        file_record["manual_adjustment"] = adjustment_record

            dit_entries = result_entries
            qc_entries = list(result.get("qc_report_entries") or [])
            artifacts = _output_records(aggregate_output_dir)
            self.payload["counts"].update(
                {
                    "completed_jobs": len(completed),
                    "failed_jobs": len(failed),
                    "dit_entries": len(dit_entries),
                    "qc_entries": len(qc_entries),
                    "patient_entries": max(0, len(dit_entries) - len(qc_entries)),
                    "html_artifacts": sum(
                        1
                        for artifact in artifacts
                        if str(artifact.get("relative_path") or "")
                        .lower()
                        .endswith(".html")
                    ),
                    "workbook_artifacts": sum(
                        1
                        for artifact in artifacts
                        if str(artifact.get("relative_path") or "")
                        .lower()
                        .endswith(".xlsx")
                    ),
                    "ladder_review_cases": int(
                        (review_gate or {}).get("review_case_count") or 0
                    ),
                }
            )
            self.payload["outputs"] = {
                "aggregate_output_dir": (
                    str(aggregate_output_dir.resolve())
                    if aggregate_output_dir is not None
                    else None
                ),
                "review_bundle": {
                    key: str(value)
                    for key, value in (review_gate or {}).items()
                    if key in {"cases_path", "summary_path"} and value
                },
                "artifacts": artifacts,
            }
            if failed:
                self.payload["status"] = "completed_with_errors"
            elif bool(result.get("dit_reports_blocked")):
                self.payload["status"] = "awaiting_ladder_review"
            else:
                self.payload["status"] = "completed"
            self.payload["completed_at_utc"] = _utc_now()
            self._write()


def load_run_manifest(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("schema_version") != RUN_MANIFEST_SCHEMA:
        raise ValueError(f"Unsupported run manifest schema: {payload.get('schema_version')}")
    return payload


def jobs_from_run_manifest(path: Path) -> list[dict[str, Any]]:
    """Rebuild original job membership for an explicit recovery rerun."""
    payload = load_run_manifest(path)
    jobs: list[dict[str, Any]] = []
    for job in payload.get("jobs") or []:
        jobs.append(
            {
                "name": str(job.get("name") or ""),
                "type": str(job.get("type") or "pipeline"),
                "path": Path(str(job["source_dir"])) if job.get("source_dir") else None,
                "files": [
                    Path(str(file_record["path"]))
                    for file_record in job.get("files") or []
                    if file_record.get("path")
                ],
            }
        )
    return jobs


def record_report_finalization(
    path: Path,
    *,
    aggregate_output_dir: Path | None,
    validation: Mapping[str, Any],
) -> dict[str, Any]:
    """Attach post-review report rebuilding evidence to an existing run."""
    payload = load_run_manifest(path)
    payload["report_finalization"] = {
        **dict(validation),
        "recorded_at_utc": _utc_now(),
        "aggregate_output_dir": (
            str(aggregate_output_dir.resolve())
            if aggregate_output_dir is not None
            else None
        ),
        "artifacts": _output_records(aggregate_output_dir),
    }
    payload["status"] = (
        "completed" if bool(validation.get("passed")) else "finalization_failed"
    )
    payload["updated_at_utc"] = _utc_now()
    _atomic_write_json(path, payload)
    return payload
