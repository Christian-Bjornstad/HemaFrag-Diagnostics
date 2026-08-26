"""Labeling session model — load tracking Excel, iterate samples, write
labels back. Used by the in-app Labeling tab so the chemist can view
plots and assign clonality labels without leaving the GUI.

The chemist's daily loop:
1. Open tracking Excel (the one the batch produces)
2. For each sample row: see the FSA plot → press a number key → label
3. Save back to Excel

Label set mirrors ``ANNOTATION_CLASSES_ORDER`` from ml_training — same
Norwegian labels, same order, so ML training and labeling speak the
same vocabulary.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Re-export the canonical label order so the GUI and the trainer never
# drift. These are Norwegian — the chemist's working language.
from core.analyses.clonality.ml_training import ANNOTATION_CLASSES_ORDER
from core.analyses.clonality.ml_training import normalize_annotation_label
from core.analyses.clonality.ml_data_contract import (
    CHEMIST_LABEL_COLUMN,
    load_tracking_run_table,
)
from core.analyses.clonality.interpretation_units import (
    CHANNEL_CHEMIST_LABEL_COLUMNS,
    InterpretationUnit,
    channel_labels_from_row,
    interpretation_units_for_assay,
)

# Keyboard shortcut map — number key → label string.
# Mirrors the order in ANNOTATION_CLASSES_ORDER.
LABEL_KEYS: dict[str, str] = {
    "1": "monoklonal",
    "2": "monoklonal_pa_poly",
    "3": "polyklonal",
    "4": "oligoklonal",
    "5": "irregulaer",
    "6": "lite_pcr_produkt",
    "7": "intet_pcr_produkt",
    "8": "qc_teknisk_fail",
    "9": "usikker_review",
}

# Reverse lookup: label → shortcut key.
LABEL_TO_KEY: dict[str, str] = {v: k for k, v in LABEL_KEYS.items()}

# Keep the chemist annotation separate from the rule-based suggestion.
LABEL_COLUMN = CHEMIST_LABEL_COLUMN
EXCLUDED_LABELING_ASSAYS = {"IKZF1", "Ktr-albumin"}


@dataclass
class LabelingSample:
    """One row from the tracking Excel, ready for labeling."""
    index: int  # row position in the original DataFrame
    dit: str
    assay: str
    well: str
    file_name: str
    source_run_dir: str
    current_label: str = ""
    identity_key: str = ""
    sample_kind: str = ""
    group: str = ""
    rule_suggestion: str = ""
    rule_review_needed: bool = False
    tracking_sheet: str = ""
    tracking_row_number: int = 0
    interpretation_units: tuple[InterpretationUnit, ...] = ()
    channel_labels: dict[str, str] = field(default_factory=dict)
    original_channel_labels: dict[str, str] = field(default_factory=dict)
    legacy_label: str = ""

    @property
    def is_labeled(self) -> bool:
        if not self.interpretation_units:
            return bool(self.current_label and self.current_label.strip())
        return all(
            bool(self.channel_labels.get(unit.channel, "").strip())
            for unit in self.interpretation_units
        )

    @property
    def labeled_unit_count(self) -> int:
        if not self.interpretation_units:
            return int(bool(self.current_label.strip()))
        return sum(
            bool(self.channel_labels.get(unit.channel, "").strip())
            for unit in self.interpretation_units
        )

    def label_for_channel(self, channel: str) -> str:
        return self.channel_labels.get(str(channel).strip().upper(), "")


@dataclass
class LabelingSession:
    """In-memory model: load Excel → iterate → label → save back.

    Does NOT touch the GUI — this is pure data. The tab widget wraps it.
    """
    excel_path: str
    include_controls: bool = False
    include_size_ladders: bool = False
    samples: list[LabelingSample] = field(default_factory=list)
    _df: pd.DataFrame | None = None
    _dirty: bool = False
    _primary_sheet: str = ""
    _available_sheets: tuple[str, ...] = field(default_factory=tuple)

    def load(self) -> None:
        """Load tracked injections from current or legacy workbook sheets."""
        table = load_tracking_run_table(
            self.excel_path,
            include_controls=self.include_controls,
            include_size_ladders=self.include_size_ladders,
        )
        df = table.frame
        self._primary_sheet = table.primary_sheet
        self._available_sheets = table.available_sheets
        for label_column in (
            LABEL_COLUMN,
            *CHANNEL_CHEMIST_LABEL_COLUMNS,
        ):
            if label_column in df.columns:
                df[label_column] = (
                    df[label_column]
                    .where(pd.notna(df[label_column]), "")
                    .map(normalize_annotation_label)
                    .astype(object)
                )
            else:
                df[label_column] = pd.Series(
                    [""] * len(df),
                    index=df.index,
                    dtype=object,
                )
        if "Assay" in df.columns:
            assay = df["Assay"].fillna("").astype(str).str.strip()
            df = df.loc[~assay.isin(EXCLUDED_LABELING_ASSAYS)].reset_index(drop=True)
        self._df = df

        self.samples = []
        for idx, row in df.iterrows():
            def _str(col, _row=row):
                v = _row.get(col, "")
                return str(v) if pd.notna(v) else ""
            assay = _str("Assay")
            units = interpretation_units_for_assay(assay)
            channel_labels = {
                channel: normalize_annotation_label(label)
                for channel, label in channel_labels_from_row(row, assay).items()
            }
            primary_label = (
                channel_labels.get(units[0].channel, "")
                if units
                else _str(LABEL_COLUMN)
            )
            sample = LabelingSample(
                index=int(idx) if pd.notna(idx) else len(self.samples),
                dit=_str("DIT"),
                assay=assay,
                well=_str("Well"),
                file_name=_str("File"),
                source_run_dir=_str("SourceRunDir"),
                current_label=primary_label,
                identity_key=_str("IdentityKey"),
                sample_kind=_str("SampleKind"),
                group=_str("Group"),
                rule_suggestion=_str("ClonalitySuggestion"),
                rule_review_needed=_as_bool(row.get("ClonalityReviewNeeded", False)),
                tracking_sheet=_str("_TrackingSheet"),
                tracking_row_number=int(row.get("_TrackingRowNumber", idx + 2) or idx + 2),
                interpretation_units=units,
                channel_labels=dict(channel_labels),
                original_channel_labels=dict(channel_labels),
                legacy_label=_str(LABEL_COLUMN),
            )
            self.samples.append(sample)

        logger.info("Loaded %d samples from %s", len(self.samples), self.excel_path)

    @property
    def total_count(self) -> int:
        return len(self.samples)

    @property
    def labeled_count(self) -> int:
        return sum(1 for s in self.samples if s.is_labeled)

    @property
    def unlabeled_count(self) -> int:
        return self.total_count - self.labeled_count

    @property
    def total_unit_count(self) -> int:
        return sum(max(1, len(sample.interpretation_units)) for sample in self.samples)

    @property
    def labeled_unit_count(self) -> int:
        return sum(sample.labeled_unit_count for sample in self.samples)

    def label_sample(
        self,
        sample_index: int,
        label: str,
        *,
        channel: str | None = None,
    ) -> None:
        """Set the label on sample at ``sample_index`` (0-based in ``self.samples``)."""
        if not (0 <= sample_index < len(self.samples)):
            raise IndexError(f"sample_index {sample_index} out of range (0-{len(self.samples) - 1})")
        label = normalize_annotation_label(label)
        if label and label not in ANNOTATION_CLASSES_ORDER:
            raise ValueError(
                f"Unknown label '{label}'. Valid: {list(ANNOTATION_CLASSES_ORDER)}"
            )
        sample = self.samples[sample_index]
        if sample.interpretation_units:
            selected = str(channel or sample.interpretation_units[0].channel).upper()
            valid_channels = {
                unit.channel for unit in sample.interpretation_units
            }
            if selected not in valid_channels:
                raise ValueError(
                    f"Channel {selected!r} is not configured for {sample.assay}"
                )
            sample.channel_labels[selected] = label
            sample.current_label = sample.channel_labels.get(
                sample.interpretation_units[0].channel,
                "",
            )
        else:
            sample.current_label = label
        self._dirty = True

    def label_all_channels(self, sample_index: int, label: str) -> int:
        if not (0 <= sample_index < len(self.samples)):
            raise IndexError(
                f"sample_index {sample_index} out of range (0-{len(self.samples) - 1})"
            )
        sample = self.samples[sample_index]
        channels = [unit.channel for unit in sample.interpretation_units]
        if not channels:
            self.label_sample(sample_index, label)
            return 1
        for channel in channels:
            self.label_sample(sample_index, label, channel=channel)
        return len(channels)

    def label_parallel_group(
        self,
        sample_index: int,
        label: str,
        *,
        channel: str | None = None,
    ) -> int:
        """Apply one chemist label to all rows for the same DIT+assay pair."""
        indices = self.parallel_indices_for(sample_index)
        for index in indices:
            self.label_sample(index, label, channel=channel)
        return len(indices)

    def clear_label(self, sample_index: int, *, channel: str | None = None) -> None:
        """Remove the label from a sample."""
        self.label_sample(sample_index, "", channel=channel)

    def save_to_excel(self) -> int:
        """Write current labels back to the Excel. Returns number of labels written."""
        if self._df is None:
            raise RuntimeError("No Excel loaded — call load() first")

        changed: list[tuple[LabelingSample, str, str]] = []
        for sample in self.samples:
            if 0 <= sample.index < len(self._df):
                if sample.interpretation_units:
                    for unit in sample.interpretation_units:
                        new_val = sample.channel_labels.get(unit.channel, "")
                        old_val = sample.original_channel_labels.get(unit.channel, "")
                        if old_val == new_val:
                            continue
                        if unit.label_column not in self._df.columns:
                            self._df[unit.label_column] = ""
                        self._df.at[
                            self._df.index[sample.index],
                            unit.label_column,
                        ] = new_val
                        changed.append((sample, unit.label_column, new_val))
                        if len(sample.interpretation_units) == 1:
                            if LABEL_COLUMN not in self._df.columns:
                                self._df[LABEL_COLUMN] = ""
                            self._df.at[
                                self._df.index[sample.index],
                                LABEL_COLUMN,
                            ] = new_val
                else:
                    old_val = sample.legacy_label
                    new_val = sample.current_label
                    if old_val != new_val:
                        if LABEL_COLUMN not in self._df.columns:
                            self._df[LABEL_COLUMN] = ""
                        self._df.at[
                            self._df.index[sample.index],
                            LABEL_COLUMN,
                        ] = new_val
                        changed.append((sample, LABEL_COLUMN, new_val))

        if changed:
            self._write_changed_labels(changed)
            for sample in self.samples:
                sample.original_channel_labels = dict(sample.channel_labels)
                if len(sample.interpretation_units) <= 1:
                    sample.legacy_label = sample.current_label

        self._dirty = False
        logger.info("Wrote %d labels to %s", len(changed), self.excel_path)
        return len(changed)

    def _write_changed_labels(
        self,
        changed: list[tuple[LabelingSample, str, str]],
    ) -> None:
        """Update label cells in-place so workbook formatting is preserved."""
        wb = load_workbook(self.excel_path)
        try:
            samples = [sample for sample, _column, _value in changed]
            target_names = {
                sample.tracking_sheet
                for sample in samples
                if sample.tracking_sheet
            }
            target_names.add(self._primary_sheet)
            if self._primary_sheet == "Runs":
                target_names.update({"Patient_Runs", "Control_Runs"})

            for sheet_name in sorted(target_names):
                if not sheet_name or sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = {
                    str(cell.value or "").strip(): cell.column
                    for cell in ws[1]
                    if str(cell.value or "").strip()
                }
                identity_col = headers.get("IdentityKey")

                identity_rows: dict[str, list[int]] = {}
                if identity_col is not None:
                    for row_number in range(2, ws.max_row + 1):
                        identity = str(ws.cell(row_number, identity_col).value or "").strip()
                        if identity:
                            identity_rows.setdefault(identity, []).append(row_number)

                for sample, label_column, new_value in changed:
                    label_col = headers.get(label_column)
                    if label_col is None:
                        label_col = ws.max_column + 1
                        ws.cell(1, label_col, label_column)
                        headers[label_column] = label_col
                    row_numbers = identity_rows.get(sample.identity_key, []) if sample.identity_key else []
                    if not row_numbers and sheet_name == sample.tracking_sheet and sample.tracking_row_number >= 2:
                        row_numbers = [sample.tracking_row_number]
                    for row_number in row_numbers:
                        ws.cell(row_number, label_col, new_value)
                        if len(sample.interpretation_units) == 1:
                            legacy_col = headers.get(LABEL_COLUMN)
                            if legacy_col is None:
                                legacy_col = ws.max_column + 1
                                ws.cell(1, legacy_col, LABEL_COLUMN)
                                headers[LABEL_COLUMN] = legacy_col
                            ws.cell(row_number, legacy_col, new_value)
            wb.save(self.excel_path)
        finally:
            wb.close()

    def filter_unlabeled(self) -> list[int]:
        """Return the indices (into ``self.samples``) of unlabeled samples."""
        return [i for i, s in enumerate(self.samples) if not s.is_labeled]

    def filter_review_needed(self) -> list[int]:
        """Return indices whose rule interpretation requires chemist review."""
        return [i for i, sample in enumerate(self.samples) if sample.rule_review_needed]

    def parallel_indices_for(self, sample_index: int) -> list[int]:
        """Return row indices for the same patient DIT and assay as a sample."""
        if not (0 <= sample_index < len(self.samples)):
            return []
        sample = self.samples[sample_index]
        dit = sample.dit.strip()
        assay = sample.assay.strip()
        if not dit or not assay:
            return [sample_index]
        indices = [
            index
            for index, candidate in enumerate(self.samples)
            if candidate.dit.strip() == dit and candidate.assay.strip() == assay
        ]
        return sorted(indices, key=lambda index: (self.samples[index].well, self.samples[index].file_name))

    def fsa_path_for(self, sample_index: int, fsa_root: str) -> Path | None:
        """Resolve the FSA file path for a sample, given the FSA root dir."""
        if not (0 <= sample_index < len(self.samples)):
            return None
        sample = self.samples[sample_index]
        if not sample.file_name:
            return None
        root = Path(fsa_root)
        if sample.source_run_dir:
            # source_run_dir is typically a folder name like "2025_01_15_FR1_..."
            candidate = root / sample.source_run_dir / sample.file_name
            if candidate.exists():
                return candidate
        # Try without source_run_dir (flat structure)
        candidate2 = root / sample.file_name
        if candidate2.exists():
            return candidate2
        return None


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if pd.isna(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}
