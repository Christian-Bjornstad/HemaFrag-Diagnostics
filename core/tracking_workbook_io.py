"""Stable, formula-friendly writes for operator tracking workbooks."""
from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _excel_value(value):
    try:
        missing = pd.isna(value)
        if not hasattr(missing, "__len__") and bool(missing):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except (AttributeError, ValueError):
            pass
    return value


def _table_name(sheet_name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)
    if not safe or safe[0].isdigit():
        safe = f"T_{safe}"
    return f"HemaFrag_{safe}"[:255]


def _headers(ws) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): int(cell.column)
        for cell in ws[1]
        if str(cell.value or "").strip()
    }


def _ensure_headers(ws, columns: Sequence[str]) -> dict[str, int]:
    headers = _headers(ws)
    for column in columns:
        name = str(column)
        if name in headers:
            continue
        index = ws.max_column + 1 if headers else 1
        ws.cell(1, index, name)
        headers[name] = index
    return headers


def _row_key(ws, row: int, headers: dict[str, int], key_columns: Sequence[str]):
    values = tuple(str(ws.cell(row, headers[column]).value or "") for column in key_columns)
    return values if all(values) else None


def _record_key(record: dict, key_columns: Sequence[str]):
    values = tuple(str(record.get(column) or "") for column in key_columns)
    return values if all(values) else None


def _copy_formula_columns(ws, source_row: int, target_row: int, generated_columns: set[str]) -> None:
    if source_row < 2 or target_row <= source_row:
        return
    headers = _headers(ws)
    for header, column in headers.items():
        if header in generated_columns:
            continue
        source = ws.cell(source_row, column)
        if not isinstance(source.value, str) or not source.value.startswith("="):
            continue
        target = ws.cell(target_row, column)
        try:
            target.value = Translator(
                source.value,
                origin=source.coordinate,
            ).translate_formula(target.coordinate)
        except Exception:
            target.value = source.value
        target.number_format = source.number_format


def _refresh_table(ws) -> None:
    if not _headers(ws):
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
    tables = list(ws.tables.values())
    if tables:
        tables[0].ref = ref
        return
    table = Table(displayName=_table_name(ws.title), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def upsert_frame(
    workbook,
    sheet_name: str,
    frame: pd.DataFrame,
    *,
    key_columns: Sequence[str],
    remove_missing: bool = False,
) -> None:
    """Update matching rows in place and append only unseen identities."""
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    columns = [str(column) for column in frame.columns]
    headers = _ensure_headers(ws, columns)
    generated_columns = set(columns)
    if remove_missing:
        desired_keys = {
            key
            for record in frame.to_dict(orient="records")
            if (key := _record_key(record, key_columns)) is not None
        }
        for row in range(ws.max_row, 1, -1):
            key = _row_key(ws, row, headers, key_columns)
            if key is not None and key not in desired_keys:
                ws.delete_rows(row)
    existing: dict[tuple[str, ...], int] = {}
    for row in range(2, ws.max_row + 1):
        key = _row_key(ws, row, headers, key_columns)
        if key is not None:
            existing[key] = row

    for record in frame.to_dict(orient="records"):
        key = _record_key(record, key_columns)
        row = existing.get(key) if key is not None else None
        if row is None:
            previous_row = ws.max_row
            row = max(2, previous_row + 1)
            _copy_formula_columns(ws, previous_row, row, generated_columns)
            if key is not None:
                existing[key] = row
        for column in columns:
            ws.cell(row, headers[column], _excel_value(record.get(column)))
    _refresh_table(ws)


def replace_frame(
    workbook,
    sheet_name: str,
    frame: pd.DataFrame,
) -> None:
    """Refresh a derived sheet while retaining its worksheet identity."""
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    columns = [str(column) for column in frame.columns]
    existing_headers = _headers(ws)
    custom_columns = [
        header for header in existing_headers if header not in columns
    ]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    headers = _ensure_headers(ws, columns)
    for row_index, record in enumerate(frame.to_dict(orient="records"), start=2):
        for column in columns:
            ws.cell(row_index, headers[column], _excel_value(record.get(column)))
    for header in custom_columns:
        column = existing_headers[header]
        if not ws.cell(1, column).value:
            ws.cell(1, column, header)
    _refresh_table(ws)


def write_tracking_frames(
    workbook_path: Path,
    frames: Iterable[
        tuple[str, pd.DataFrame, Sequence[str] | None]
        | tuple[str, pd.DataFrame, Sequence[str] | None, bool]
    ],
) -> None:
    """Write tracking frames to an existing or new workbook without replacing sheets."""
    path = Path(workbook_path)
    if path.exists():
        workbook = load_workbook(path, keep_links=True)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
    try:
        for item in frames:
            sheet_name, frame, keys = item[:3]
            remove_missing = bool(item[3]) if len(item) > 3 else False
            if keys:
                upsert_frame(
                    workbook,
                    sheet_name,
                    frame,
                    key_columns=keys,
                    remove_missing=remove_missing,
                )
            else:
                replace_frame(workbook, sheet_name, frame)
        workbook.save(path)
    finally:
        workbook.close()


def publish_workbook_contents(staged_path: Path, destination: Path) -> None:
    """Publish a validated workbook while retaining an existing file's identity."""
    staged = Path(staged_path)
    target = Path(destination)
    if not target.exists():
        os.replace(staged, target)
        return
    with staged.open("rb") as source, target.open("r+b") as output:
        output.seek(0)
        shutil.copyfileobj(source, output, length=1024 * 1024)
        output.truncate()
        output.flush()
        os.fsync(output.fileno())
    staged.unlink()


__all__ = [
    "publish_workbook_contents",
    "replace_frame",
    "upsert_frame",
    "write_tracking_frames",
]
