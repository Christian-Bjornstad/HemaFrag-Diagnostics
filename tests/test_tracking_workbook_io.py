from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.tracking_workbook_io import (
    publish_workbook_contents,
    write_tracking_frames,
)


def test_tracking_rows_upsert_and_extend_user_formula_columns(tmp_path):
    path = tmp_path / "tracking.xlsx"
    first = pd.DataFrame(
        [{"IdentityKey": "one", "Value": 2}]
    )
    write_tracking_frames(path, (("Runs", first, ("IdentityKey",)),))

    workbook = load_workbook(path)
    sheet = workbook["Runs"]
    sheet.cell(1, 3, "DoubleValue")
    sheet.cell(2, 3, "=B2*2")
    workbook.save(path)
    workbook.close()

    updated = pd.DataFrame(
        [
            {"IdentityKey": "one", "Value": 3},
            {"IdentityKey": "two", "Value": 5},
        ]
    )
    write_tracking_frames(path, (("Runs", updated, ("IdentityKey",)),))

    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Runs"]
    assert sheet["B2"].value == 3
    assert sheet["C2"].value == "=B2*2"
    assert sheet["B3"].value == 5
    assert sheet["C3"].value == "=B3*2"
    assert list(sheet.tables.values())[0].ref == "A1:C3"
    workbook.close()


def test_publishing_existing_workbook_keeps_destination_entry(tmp_path, monkeypatch):
    destination = tmp_path / "tracking.xlsx"
    destination.write_bytes(b"old")
    staged = tmp_path / "staged.xlsx"
    staged.write_bytes(b"new workbook bytes")

    monkeypatch.setattr(
        "core.tracking_workbook_io.os.replace",
        lambda *_args: (_ for _ in ()).throw(AssertionError("must not replace")),
    )

    publish_workbook_contents(staged, destination)

    assert destination.read_bytes() == b"new workbook bytes"
    assert not staged.exists()
