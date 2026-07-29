from __future__ import annotations

import copy
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from config import APP_SETTINGS
from core.analyses.clonality.ml_data_contract import CHEMIST_LABEL_COLUMN
from core.analyses.clonality.tracking_excel import update_clonality_tracking_workbook
from core.labeling.labeling_session import LabelingSession


def _entry(file_name: str, *, assay: str = "FR1", dit: str = "") -> dict:
    return {
        "fsa": None,
        "file_name": file_name,
        "assay": assay,
        "dit": dit,
        "group": "B",
        "ladder": "ROX400HD",
        "ladder_qc_status": "ok",
        "ladder_fit_strategy": "linear",
        "ladder_expected_step_count": 16,
        "ladder_fitted_step_count": 16,
        "ladder_r2": 0.9999,
        "ladder_linear_r2": 0.9999,
        "ladder_linear_mean_residual_bp": 0.2,
        "ladder_linear_max_residual_bp": 0.7,
        "ladder_max_curvature": 0.0,
    }


class ClonalityTrackingOutputTests(unittest.TestCase):
    def setUp(self) -> None:
        self._settings = copy.deepcopy(APP_SETTINGS)

    def tearDown(self) -> None:
        APP_SETTINGS.clear()
        APP_SETTINGS.update(self._settings)

    def test_tracking_workbook_writes_patient_control_and_dashboard_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            entries = [
                _entry("26OUM00001_FR1__220526_A01_H9TEST01.fsa", dit="26OUM00001"),
                _entry("PK_FR1__220526_E08_H9TEST01.fsa"),
            ]
            marker = {
                "name": "FR1_PK",
                "kind": "sample",
                "channel": "DATA1",
                "expected_bp": 100.0,
                "window_bp": 3.0,
            }

            with patch("core.analyses.clonality.tracking_excel.markers_for_entry", return_value=[marker]):
                update_clonality_tracking_workbook(workbook, entries)

            with pd.ExcelFile(workbook, engine="openpyxl") as xls:
                sheets = xls.sheet_names
            self.assertIn("Runs", sheets)
            self.assertIn("Patient_Runs", sheets)
            self.assertIn("Control_Runs", sheets)
            self.assertIn("PK_Peaks", sheets)
            self.assertIn("Dashboard", sheets)
            self.assertEqual(
                sheets,
                ["Dashboard", "Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"],
            )

            patients = pd.read_excel(workbook, sheet_name="Patient_Runs", engine="openpyxl")
            controls = pd.read_excel(workbook, sheet_name="Control_Runs", engine="openpyxl")
            peaks = pd.read_excel(workbook, sheet_name="PK_Peaks", engine="openpyxl")
            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            self.assertEqual(len(patients), 1)
            self.assertEqual(len(controls), 1)
            self.assertEqual(controls.iloc[0]["Control"], "PK")
            self.assertEqual(len(peaks), 1)
            self.assertIn("ManualAdjustmentUsed", runs.columns)
            for removed in (
                "SourceFsaSha256",
                "ManualAdjustmentSha256",
                "AnalysisVersion",
                "LadderEngine",
                "LadderReasonCodes",
                "PullUpCandidate",
                "SaturationCandidate",
            ):
                self.assertNotIn(removed, runs.columns)

            wb = load_workbook(workbook, data_only=False)
            formulas = [
                cell.value
                for sheet in wb.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            self.assertFalse(any("1048576" in formula for formula in formulas))
            self.assertFalse(
                any(
                    f"${letter}:${letter}" in formula
                    for formula in formulas
                    for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                )
            )
            dashboard = wb["Dashboard"]
            self.assertEqual(dashboard["A16"].value, "Assay")
            self.assertEqual(dashboard["A17"].value, "FR1")
            self.assertEqual(dashboard["C5"].value, 0)
            for sheet_name in ("Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"):
                sheet = wb[sheet_name]
                self.assertEqual(
                    sheet.freeze_panes,
                    "F2" if sheet_name == "PK_Peaks" else "G2",
                )
                self.assertEqual(sheet.auto_filter.ref, sheet.dimensions)
                identity_column = next(
                    cell.column_letter
                    for cell in sheet[1]
                    if cell.value == "IdentityKey"
                )
                self.assertTrue(sheet.column_dimensions[identity_column].hidden)
            wb.close()

    def test_manual_adjustment_is_tracked_but_not_counted_as_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            manual = _entry(
                "26OUM00001_FR1__220526_A01_H9TEST01.fsa",
                dit="26OUM00001",
            )
            manual["ladder_qc_status"] = "manual_adjustment"
            manual["ladder_fit_strategy"] = "manual_adjustment"
            review = _entry(
                "26OUM00002_FR1__220526_A02_H9TEST01.fsa",
                dit="26OUM00002",
            )
            review["ladder_qc_status"] = "review_required"

            update_clonality_tracking_workbook(workbook, [manual, review])

            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            manual_row = runs.loc[runs["LadderFitStrategy"].eq("manual_adjustment")].iloc[0]
            self.assertTrue(bool(manual_row["ManualAdjustmentUsed"]))
            wb = load_workbook(workbook, data_only=True)
            dashboard = wb["Dashboard"]
            self.assertEqual(dashboard["C5"].value, 1)
            self.assertEqual(dashboard["D5"].value, 1)
            self.assertEqual(dashboard["E17"].value, 1)
            self.assertEqual(dashboard["F17"].value, 1)
            wb.close()

    def test_tracking_workbook_writes_ml_columns_when_present(self) -> None:
        import math
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            patient_entry = _entry(
                "26OUM00001_FR1__220526_A01_H9TEST01.fsa", dit="26OUM00001"
            )
            # Stamp ML fields onto the entry
            patient_entry["ClonalitySuggestion"] = "monoklonal"
            patient_entry["ClonalityConfidence"] = 0.93
            patient_entry["ClonalityReviewNeeded"] = False
            patient_entry["ClonalityMLSuggestion"] = "monoklonal"
            patient_entry["ClonalityMLConfidence"] = 0.86
            patient_entry["ClonalityMLThreshold"] = 0.8
            patient_entry["ClonalityMLReviewNeeded"] = False
            patient_entry["ClonalityMLEvidence"] = "rule_ml_agree"
            patient_entry["ClonalityMLModelVersion"] = "ml_training_pipeline_v9"
            entries = [
                patient_entry,
                # Control entry does NOT carry ML fields (chemist usually
                # only stamps ML onto patient samples). The Tracking
                # Excel writer must not invent values for absent ML keys.
                _entry("PK_FR1__220526_E08_H9TEST01.fsa"),
            ]
            marker = {
                "name": "FR1_PK",
                "kind": "sample",
                "channel": "DATA1",
                "expected_bp": 100.0,
                "window_bp": 3.0,
            }
            with patch(
                "core.analyses.clonality.tracking_excel.markers_for_entry",
                return_value=[marker],
            ):
                update_clonality_tracking_workbook(workbook, entries)
            patients = pd.read_excel(workbook, sheet_name="Patient_Runs", engine="openpyxl")
            self.assertEqual(len(patients), 1)
            # All ML columns are present and round-trip
            self.assertIn("ClonalityMLSuggestion", patients.columns)
            self.assertIn("ClonalityMLConfidence", patients.columns)
            self.assertIn("ClonalityMLThreshold", patients.columns)
            self.assertIn("ClonalityMLReviewNeeded", patients.columns)
            self.assertIn("ClonalityMLEvidence", patients.columns)
            self.assertIn("ClonalityMLModelVersion", patients.columns)
            self.assertEqual(patients.iloc[0]["ClonalityMLSuggestion"], "monoklonal")
            self.assertEqual(float(patients.iloc[0]["ClonalityMLConfidence"]), 0.86)
            self.assertEqual(float(patients.iloc[0]["ClonalityMLThreshold"]), 0.8)
            self.assertEqual(patients.iloc[0]["ClonalityMLEvidence"], "rule_ml_agree")
            # Control row: nothing was stamped ⇒ empty cell (Excel NaN read is fine)
            controls = pd.read_excel(workbook, sheet_name="Control_Runs", engine="openpyxl")
            self.assertEqual(len(controls), 1)
            for col in (
                "ClonalityMLSuggestion",
                "ClonalityMLConfidence",
                "ClonalityMLThreshold",
                "ClonalityMLReviewNeeded",
                "ClonalityMLEvidence",
                "ClonalityMLModelVersion",
            ):
                val = controls.iloc[0][col]
                # Empty string OR NaN both acceptable — both signal "no ML".
                if isinstance(val, float):
                    self.assertTrue(math.isnan(val), f"{col} expected empty, got {val!r}")
                else:
                    self.assertEqual(val, "")

    def test_batch_refresh_preserves_chemist_label(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            entry = _entry(
                "26OUM00001_FR1__220526_A01_H9TEST01.fsa",
                dit="26OUM00001",
            )
            update_clonality_tracking_workbook(workbook, [entry])

            session = LabelingSession(excel_path=str(workbook))
            session.load()
            session.label_sample(0, "monoklonal")
            self.assertEqual(session.save_to_excel(), 1)

            update_clonality_tracking_workbook(workbook, [entry])

            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            self.assertEqual(runs.iloc[0][CHEMIST_LABEL_COLUMN], "monoklonal")
            self.assertEqual(patients.iloc[0][CHEMIST_LABEL_COLUMN], "monoklonal")

    def test_tracking_workbook_derives_missing_dit_case_insensitively(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            entry = _entry(
                "25oum12345_FR1__220526_A01_H9TEST01.fsa",
                dit="",
            )

            update_clonality_tracking_workbook(workbook, [entry])

            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            self.assertEqual(runs.iloc[0]["DIT"], "25OUM12345")

    def test_tracking_workbook_keeps_unassigned_injection_out_of_patient_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "Clonality_Tracking.xlsx"
            entry = _entry("IKZF1_unknown__220526_A01_H9TEST01.fsa", dit="")

            update_clonality_tracking_workbook(workbook, [entry])

            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            self.assertEqual(runs.iloc[0]["SampleKind"], "unassigned")
            self.assertTrue(patients.empty)

    def test_aggregated_batch_uses_one_local_tracking_workbook_and_global_dashboard(self) -> None:
        import core.batch as batch

        with tempfile.TemporaryDirectory() as tmp:
            output_root = Path(tmp) / "run"
            global_path = Path(tmp) / "HemaFrag_Clonality_All_Runs.xlsx"
            APP_SETTINGS["active_analysis"] = "clonality"
            APP_SETTINGS.setdefault("analyses", {}).setdefault("clonality", {}).setdefault("batch", {})
            APP_SETTINGS["analyses"]["clonality"]["batch"].update(
                {
                    "global_tracking_excel_path": str(global_path),
                    "ladder_review_gate": {"enabled": False},
                }
            )

            patient_entry = _entry("26OUM00001_FR1__220526_A01_H9TEST01.fsa", dit="26OUM00001")
            qc_entry = _entry("PK_FR1__220526_E08_H9TEST01.fsa")
            jobs = [
                {"name": "26OUM00001", "type": "pipeline", "path": None, "files": [Path("patient.fsa")]},
                {"name": "QC", "type": "qc", "path": None, "files": [Path("pk.fsa")]},
            ]
            qc_kwargs = {}

            def fake_qc_job(**kwargs):
                qc_kwargs.update(kwargs)
                return None, [qc_entry]

            with (
                patch.object(batch, "run_pipeline_job_collect", return_value=[patient_entry]),
                patch.object(batch, "run_qc_job", side_effect=fake_qc_job),
                patch("core.html_reports.build_dit_html_reports", return_value=None),
            ):
                result = batch.run_batch_jobs(
                    jobs=jobs,
                    output_base=output_root,
                    out_folder_tmpl="ASSAY_REPORTS",
                    outfile_html_tmpl="QC_REPORT_{name}.html",
                    excel_name_tmpl="HemaFrag_QC_Trends.xlsx",
                    pipeline_scope="all",
                    assay_filter="",
                    aggregate_dit_reports=True,
                    continue_on_error=False,
                    aggregate_outdir_name="reports_test",
                )

            self.assertEqual(result["failed_jobs"], [])
            manifest_path = Path(result["run_manifest_path"])
            self.assertTrue(manifest_path.is_file())
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["status"], "completed")
            self.assertEqual(manifest["counts"]["expected_jobs"], 2)
            self.assertEqual(manifest["counts"]["dit_entries"], 2)
            self.assertEqual(manifest["counts"]["qc_entries"], 1)
            self.assertFalse(qc_kwargs["update_qc_trends"])
            self.assertFalse((output_root / "ASSAY_REPORTS").exists())
            self.assertFalse((output_root / "HemaFrag_QC_Trends.xlsx").exists())

            local_workbook = output_root / "reports_test" / "Clonality_Tracking.xlsx"
            self.assertTrue(local_workbook.exists())
            self.assertTrue(global_path.exists())

            local_runs = pd.read_excel(local_workbook, sheet_name="Runs", engine="openpyxl")
            global_runs = pd.read_excel(global_path, sheet_name="Runs", engine="openpyxl")
            self.assertEqual(set(local_runs["SampleKind"]), {"patient", "control"})
            self.assertEqual(set(global_runs["SampleKind"]), {"patient", "control"})


if __name__ == "__main__":
    unittest.main()
