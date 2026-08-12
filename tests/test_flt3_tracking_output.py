from __future__ import annotations

import copy
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config import APP_SETTINGS
from core.analyses.flt3.qc_tracker import update_flt3_npm1_qc_tracker_workbook
from scripts.run_flt3_liz500_qc_all_injections import (
    QC_OUTPUT_COLUMNS,
    RAW_METADATA_COLUMNS,
    _update_global_rox500_workbook,
)


def _entry(file_name: str, *, specimen_id: str = "26OUM00001", assay: str = "FLT3-ITD") -> dict:
    return {
        "fsa": None,
        "file_name": file_name,
        "source_run_dir": "flt3_run",
        "dit": "26OUM00001" if specimen_id.startswith("26OUM") else "",
        "assay": assay,
        "specimen_id": specimen_id,
        "analysis_type": "",
        "group": "sample",
        "run_date": "2026-05-26",
        "well_id": "A01",
        "ladder": "GS500ROX",
        "ladder_qc_status": "ok",
        "ladder_fit_strategy": "linear",
        "ladder_expected_step_count": 16,
        "ladder_fitted_step_count": 16,
        "ladder_r2": 0.9999,
        "peak_qc_status": "ok",
        "primary_peak_channel": "DATA1",
        "peaks_by_channel": {
            "DATA1": pd.DataFrame(
                [
                    {"label": "WT", "basepairs": 329.0, "peaks": 1200.0, "area": 5000.0, "keep": True},
                    {"label": "MUT", "basepairs": 360.0, "peaks": 200.0, "area": 300.0, "keep": True},
                ]
            )
        },
        "ratio": 0.0,
    }


class Flt3TrackingOutputTests(unittest.TestCase):
    def setUp(self) -> None:
        self._settings = copy.deepcopy(APP_SETTINGS)

    def tearDown(self) -> None:
        APP_SETTINGS.clear()
        APP_SETTINGS.update(self._settings)

    def test_flt3_tracking_workbook_writes_split_dashboard_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "FLT3_Tracking.xlsx"
            update_flt3_npm1_qc_tracker_workbook(
                workbook,
                [
                    _entry("26OUM00001_ITD__260526_A01_H9TEST01.fsa"),
                    _entry("IVS-0000_ITD__260526_B01_H9TEST01.fsa", specimen_id="IVS-0000"),
                    _entry(
                        "IVS-P001_D835__260526_C01_H9TEST01.fsa",
                        specimen_id="IVS-P001",
                        assay="FLT3-D835",
                    ),
                ],
            )

            with pd.ExcelFile(workbook, engine="openpyxl") as xls:
                sheets = xls.sheet_names
            self.assertIn("Runs", sheets)
            self.assertIn("Patient_Runs", sheets)
            self.assertIn("Control_Runs", sheets)
            self.assertIn("PK_Peaks", sheets)
            self.assertIn("Dashboard", sheets)
            self.assertEqual(
                sheets,
                ["Dashboard", "Runs", "Patient_Runs", "Control_Runs", "PK_Peaks"],
            )

            patients = pd.read_excel(workbook, sheet_name="Patient_Runs", engine="openpyxl")
            controls = pd.read_excel(workbook, sheet_name="Control_Runs", engine="openpyxl")
            peaks = pd.read_excel(workbook, sheet_name="PK_Peaks", engine="openpyxl")
            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            self.assertEqual(len(patients), 1)
            self.assertEqual(len(controls), 2)
            self.assertEqual(set(controls["Control"]), {"RK", "PK"})
            self.assertEqual(len(peaks), 1)
            self.assertEqual(peaks.iloc[0]["Assay"], "FLT3-D835")
            self.assertEqual(peaks.iloc[0]["Kind"], "sample")
            self.assertEqual(peaks.iloc[0]["MarkerName"], "IVSP001_D835_128_129")
            self.assertIn("ManualAdjustmentUsed", runs.columns)
            for removed in (
                "SourceFsaSha256",
                "ManualAdjustmentSha256",
                "AnalysisVersion",
                "LadderEngine",
                "LadderReasonCodes",
                "PullUpCandidate",
                "SaturationCandidate",
            ):
                self.assertNotIn(removed, runs.columns)

            wb = load_workbook(workbook)
            dashboard = wb["Dashboard"]
            self.assertEqual(dashboard["A16"].value, "Assay")
            self.assertEqual(dashboard["A17"].value, "FLT3-D835")
            self.assertEqual(dashboard["E6"].value, 0)
            wb.close()

    def test_flt3_tracking_preserves_rox500_qc_sheets_in_shared_global_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "HemaFrag_FLT3_All_Runs.xlsx"
            row = {column: "" for column in QC_OUTPUT_COLUMNS}
            row.update(
                {
                    "File": "IVS-0000_ITD__260526_B01_H9TEST01.fsa",
                    "SourceRunDir": "flt3_run",
                    "InjectionTimeSeconds": 5,
                    "QCStatus": "PASS",
                    "LadderQC": "ok",
                    "PeakQC": "not_evaluated",
                }
            )
            raw_row = {column: "" for column in RAW_METADATA_COLUMNS}
            raw_row.update({"File": row["File"], "RunName": "flt3_run", "InjectionTimeSeconds": 5})
            _update_global_rox500_workbook(
                workbook,
                qc_df=pd.DataFrame([row], columns=QC_OUTPUT_COLUMNS),
                raw_meta_df=pd.DataFrame([raw_row], columns=RAW_METADATA_COLUMNS),
                skipped=[],
            )

            update_flt3_npm1_qc_tracker_workbook(
                workbook,
                [_entry("26OUM00001_ITD__260526_A01_H9TEST01.fsa")],
            )

            with pd.ExcelFile(workbook, engine="openpyxl") as xls:
                sheets = xls.sheet_names
            self.assertIn("Runs", sheets)
            self.assertIn("All_Analyzed_QC", sheets)
            self.assertIn("Raw_Metadata_All_FSA", sheets)

    def test_flt3_tracking_migrates_historical_qc_patient_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "HemaFrag_FLT3_All_Runs.xlsx"
            row = {column: "" for column in QC_OUTPUT_COLUMNS}
            row.update(
                {
                    "File": "26OUM00071_D835__130126_A05_H9C0U3IF.fsa",
                    "SourceRunDir": "2026_01_14_FLT3_PR",
                    "Well": "A05",
                    "SpecimenID": "26OUM00071",
                    "Assay": "FLT3-D835",
                    "Treatment": "standard",
                    "InjectionTimeSeconds": 1,
                    "RunDate": "2026-01-14",
                    "InternalLadder": "GS500ROX",
                    "LadderQC": "manual_adjustment",
                    "LadderFitStrategy": "manual_adjustment",
                    "LadderR2": 0.99999,
                    "PeakQC": "not_evaluated_ladder_only",
                    "QCStatus": "PASS",
                }
            )
            raw_row = {column: "" for column in RAW_METADATA_COLUMNS}
            raw_row.update(
                {
                    "File": row["File"],
                    "RunName": row["SourceRunDir"],
                    "InjectionTimeSeconds": 1,
                }
            )
            _update_global_rox500_workbook(
                workbook,
                qc_df=pd.DataFrame([row], columns=QC_OUTPUT_COLUMNS),
                raw_meta_df=pd.DataFrame(
                    [raw_row],
                    columns=RAW_METADATA_COLUMNS,
                ),
                skipped=[],
            )

            update_flt3_npm1_qc_tracker_workbook(workbook, [])

            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            self.assertEqual(len(patients), 1)
            self.assertEqual(patients.iloc[0]["DIT"], "26OUM00071")
            self.assertEqual(patients.iloc[0]["ResultStatus"], "qc_only")
            self.assertTrue(bool(patients.iloc[0]["ManualAdjustmentUsed"]))

    def test_flt3_tracking_records_results_and_preserves_review_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "FLT3_Tracking.xlsx"
            entry = _entry("26OUM00001_ITD__260526_A01_H9TEST01.fsa")
            peaks = entry["peaks_by_channel"]["DATA1"].copy()
            peaks["peak_id"] = ["wt-1", "mut-1"]
            entry["peaks_by_channel"]["DATA1"] = peaks
            entry["manual_ratio_selection"] = {
                "enabled": True,
                "wt": {"peak_id": "wt-1", "channel": "DATA1"},
                "mutants": [{"peak_id": "mut-1", "channel": "DATA1"}],
            }
            entry.update(
                {
                    "ratio": 0.06,
                    "ratio_numerator_area": 300.0,
                    "ratio_denominator_area": 5000.0,
                    "mutant_fraction": 300.0 / 5300.0,
                    "peak_qc_pass": True,
                }
            )

            update_flt3_npm1_qc_tracker_workbook(workbook, [entry])
            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            row = runs.iloc[0]
            self.assertEqual(row["ResultStatus"], "complete")
            self.assertAlmostEqual(float(row["WT_BP"]), 329.0)
            self.assertAlmostEqual(float(row["MutantMain_BP"]), 360.0)
            self.assertAlmostEqual(float(row["Ratio"]), 0.06)
            self.assertTrue(bool(row["PositiveCall"]))
            self.assertEqual(row["Interpretation"], "Positiv FLT3-ITD")

            wb = load_workbook(workbook)
            patient = wb["Patient_Runs"]
            headers = {
                str(cell.value): cell.column
                for cell in patient[1]
                if cell.value is not None
            }
            patient.cell(2, headers["ReviewStatus"], "Reviewed")
            patient.cell(2, headers["TrackingNote"], "Verified by operator")
            wb.save(workbook)
            wb.close()

            rerun = copy.deepcopy(entry)
            rerun["ladder_r2"] = 0.99995
            update_flt3_npm1_qc_tracker_workbook(workbook, [rerun])

            runs = pd.read_excel(workbook, sheet_name="Runs", engine="openpyxl")
            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            self.assertEqual(len(runs), 1)
            self.assertEqual(len(patients), 1)
            self.assertEqual(runs.iloc[0]["ReviewStatus"], "Reviewed")
            self.assertEqual(runs.iloc[0]["TrackingNote"], "Verified by operator")
            self.assertAlmostEqual(float(runs.iloc[0]["LadderR2"]), 0.99995)

    def test_flt3_tracking_classifies_pk_alias_as_control(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "FLT3_Tracking.xlsx"
            update_flt3_npm1_qc_tracker_workbook(
                workbook,
                [
                    _entry(
                        "PK-D835_D835__260526_C01_H9TEST01.fsa",
                        specimen_id="PK-D835",
                        assay="FLT3-D835",
                    )
                ],
            )

            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            controls = pd.read_excel(
                workbook,
                sheet_name="Control_Runs",
                engine="openpyxl",
            )
            self.assertTrue(patients.empty)
            self.assertEqual(len(controls), 1)
            self.assertEqual(controls.iloc[0]["Control"], "PK")
            self.assertEqual(controls.iloc[0]["SpecimenID"], "IVS-P001")

    def test_flt3_tracking_derives_patient_from_file_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "FLT3_Tracking.xlsx"
            entry = _entry(
                "25oum12345_D835__260526_A01_H9TEST01.fsa",
                specimen_id="",
                assay="FLT3-D835",
            )
            entry["dit"] = ""

            update_flt3_npm1_qc_tracker_workbook(workbook, [entry])

            patients = pd.read_excel(
                workbook,
                sheet_name="Patient_Runs",
                engine="openpyxl",
            )
            self.assertEqual(len(patients), 1)
            self.assertEqual(patients.iloc[0]["DIT"], "25OUM12345")
            self.assertEqual(patients.iloc[0]["SampleKind"], "patient")

    def test_rox500_global_workbook_appends_and_dedupes_qc_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "HemaFrag_FLT3_All_Runs.xlsx"
            row = {column: "" for column in QC_OUTPUT_COLUMNS}
            row.update(
                {
                    "File": "IVS-0000_ITD__260526_B01_H9TEST01.fsa",
                    "SourceRunDir": "flt3_run",
                    "Assay": "FLT3-ITD",
                    "ControlPrefix": "IVS-0000",
                    "InjectionTimeSeconds": 5,
                    "QCStatus": "PASS",
                    "LadderQC": "ok",
                    "PeakQC": "not_evaluated",
                }
            )
            raw_row = {column: "" for column in RAW_METADATA_COLUMNS}
            raw_row.update(
                {
                    "File": row["File"],
                    "RunName": "flt3_run",
                    "InjectionTimeSeconds": 5,
                }
            )
            qc_df = pd.DataFrame([row], columns=QC_OUTPUT_COLUMNS)
            raw_df = pd.DataFrame([raw_row], columns=RAW_METADATA_COLUMNS)

            _update_global_rox500_workbook(workbook, qc_df=qc_df, raw_meta_df=raw_df, skipped=[])
            _update_global_rox500_workbook(workbook, qc_df=qc_df, raw_meta_df=raw_df, skipped=[])

            all_qc = pd.read_excel(workbook, sheet_name="All_Analyzed_QC", engine="openpyxl")
            self.assertEqual(len(all_qc), 1)
            with pd.ExcelFile(workbook, engine="openpyxl") as xls:
                self.assertNotIn("Summary_By_Injection", xls.sheet_names)
                self.assertNotIn("Review_Rows", xls.sheet_names)


if __name__ == "__main__":
    unittest.main()
